#!/usr/bin/env python3
"""Repository-wide integrity validation for the AGCP v2.0.1 correction set."""

from __future__ import annotations

import argparse
import collections
import hashlib
import json
import re
import sys
import urllib.parse
from pathlib import Path
from typing import Any, Iterable

import yaml
from jsonschema import Draft202012Validator
from openpyxl import load_workbook

RELEASE_CONTEXT = {
    "repository_release_target": "v2.0.1",
    "repository_release_target_status": "UNRELEASED_ACCUMULATED_CORRECTION_SET",
    "controlling_published_baseline": "v2.0.0",
    "controlling_baseline_status": "PUBLIC_REVIEW_CONTROLLED_BASELINE",
    "baseline_date": "2026-07-30",
    "artifact_lifecycle_state": "CURRENT",
}

FINDINGS = [
    "P0-01", "P0-02", "P0-06", "P0-10",
    "P1-01", "P1-03", "P1-09", "P1-12", "P1-14", "P1-17",
    "P2-01", "P2-02", "P2-04", "P2-06",
]

EXPECTED_VERSIONS = {
    "rtm_dataset": "RTM-1.46",
    "schema_catalog": "1.0.50",
    "interface_catalog": "1.0.5",
    "registry_entry_catalog": "1.0.3",
    "implementation_profile_catalog": "1.0.3",
}

CONTROLLED_REPORTS = {
    "implementation_profile": "governance/AGCP-implementation-profile-validation.json",
    "provenance_wire": "governance/AGCP-provenance-wire-format-validation.json",
    "command_record": "governance/AGCP-command-record-separation-validation.json",
    "content_digest": "governance/AGCP-content-digest-contract-validation.json",
    "http_error_metadata": "governance/AGCP-http-error-metadata-validation.json",
    "semantic_fixtures": "governance/AGCP-semantic-fixture-validation.json",
    "normative_companions": "governance/AGCP-normative-companion-reference-validation.json",
    "release_lifecycle": "governance/AGCP-release-lifecycle-metadata-validation.json",
    "repository_synchronization": "governance/AGCP-v2.0.1-repository-synchronization-validation.json",
    "traceability_gap_closure": "governance/AGCP-traceability-gap-closure-validation.json",
    "release_payload_deduplication": "governance/AGCP-release-payload-deduplication-validation.json",
    "rtm_companion_dispositions": "governance/RTM-1.46-companion-artifact-disposition-validation.json",
    "schema_catalog": "schemas/catalog/schema-catalog-validation.json",
    "interface_traceability": "api/interface-traceability-validation.json",
    "openapi_migration": "api/AGCP-OpenAPI-v2-migration-validation.json",
    "registry_traceability": "registries/registry-entry-traceability-validation.json",
    "test_mapping": "conformance/test-mapping-validation.json",
    "fixture_synchronization": "conformance/AGCP-conformance-fixture-synchronization-validation.json",
    "if001_operation_coverage": "conformance/AGCP-if001-executable-operation-coverage-validation.json",
    "harness_error_model": "conformance/AGCP-harness-error-model-validation.json",
    "harness_request_parameters": "conformance/AGCP-harness-request-parameter-validation.json",
    "governance_compilation_activation": "conformance/AGCP-governance-compilation-activation-executable-validation.json",
}

INTEGRITY_REPORT = "governance/AGCP-v2.0.1-repository-integrity-validation.json"
SYNC_MANIFEST = "governance/AGCP-v2.0.1-repository-synchronization-manifest.json"
SYNC_REPORT = "governance/AGCP-v2.0.1-repository-synchronization-validation.json"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def json_pointer_exists(document: Any, fragment: str) -> bool:
    if not fragment or fragment == "#":
        return True
    if not fragment.startswith("#/"):
        return True
    current = document
    for token in fragment[2:].split("/"):
        token = token.replace("~1", "/").replace("~0", "~")
        if isinstance(current, list):
            try:
                current = current[int(token)]
            except (ValueError, IndexError):
                return False
        elif isinstance(current, dict) and token in current:
            current = current[token]
        else:
            return False
    return True


def walk_refs(value: Any, base_file: Path) -> Iterable[str]:
    if isinstance(value, dict):
        for key, child in value.items():
            if key == "$ref" and isinstance(child, str):
                yield child
            else:
                yield from walk_refs(child, base_file)
    elif isinstance(value, list):
        for child in value:
            yield from walk_refs(child, base_file)


def collect_source_hashes(value: Any) -> Iterable[tuple[str, str]]:
    if isinstance(value, dict):
        source_hashes = value.get("source_hashes")
        if isinstance(source_hashes, dict):
            for path, digest in source_hashes.items():
                if isinstance(path, str) and isinstance(digest, str):
                    yield path, digest
        for child in value.values():
            yield from collect_source_hashes(child)
    elif isinstance(value, list):
        for child in value:
            yield from collect_source_hashes(child)


def report_status(report: dict[str, Any]) -> str | None:
    value = report.get("status")
    if isinstance(value, str):
        return value.upper()
    return None


def add_check(checks: list[dict[str, Any]], name: str, passed: bool, detail: Any = None) -> None:
    checks.append({"check": name, "passed": passed, "detail": detail})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", default=".")
    parser.add_argument("--report", default=INTEGRITY_REPORT)
    args = parser.parse_args()

    root = Path(args.repo).resolve()
    report_path = (root / args.report).resolve()
    canonical_integrity_report = (root / INTEGRITY_REPORT).resolve()
    recursive_report_exclusions = {report_path, canonical_integrity_report}
    issues: list[str] = []
    checks: list[dict[str, Any]] = []

    # Parse every JSON and YAML artifact.
    parsed: dict[Path, Any] = {}
    json_count = 0
    yaml_count = 0
    for path in sorted(root.rglob("*")):
        if not path.is_file() or ".git" in path.parts or path.resolve() in recursive_report_exclusions:
            continue
        try:
            if path.suffix.lower() == ".json":
                parsed[path] = json.loads(path.read_text(encoding="utf-8"))
                json_count += 1
            elif path.suffix.lower() in {".yaml", ".yml"}:
                parsed[path] = yaml.safe_load(path.read_text(encoding="utf-8"))
                yaml_count += 1
        except Exception as exc:
            issues.append(f"parse:{path.relative_to(root)}:{exc}")
    add_check(checks, "all_json_and_yaml_parse", not any(x.startswith("parse:") for x in issues), {
        "json_files": json_count,
        "yaml_files": yaml_count,
    })

    # Verify every local JSON/YAML $ref, including JSON Pointer fragments.
    ref_count = 0
    ref_issues: list[str] = []
    for base_file, document in parsed.items():
        for ref in walk_refs(document, base_file):
            ref_count += 1
            if ref.startswith(("http://", "https://", "urn:")):
                continue
            path_part, separator, fragment = ref.partition("#")
            target = base_file if not path_part else (base_file.parent / urllib.parse.unquote(path_part)).resolve()
            if not target.is_file():
                ref_issues.append(f"missing:{base_file.relative_to(root)}:{ref}")
                continue
            if separator:
                target_doc = parsed.get(target)
                if target_doc is None:
                    try:
                        target_doc = json.loads(target.read_text(encoding="utf-8")) if target.suffix.lower() == ".json" else yaml.safe_load(target.read_text(encoding="utf-8"))
                    except Exception as exc:
                        ref_issues.append(f"parse-target:{base_file.relative_to(root)}:{ref}:{exc}")
                        continue
                if not json_pointer_exists(target_doc, "#" + fragment):
                    ref_issues.append(f"fragment:{base_file.relative_to(root)}:{ref}")
    issues.extend("ref:" + item for item in ref_issues)
    add_check(checks, "all_local_schema_and_openapi_refs_resolve", not ref_issues, {"references_checked": ref_count})

    # Validate every active schema and the additional controlled profile/IF-002 schemas against Draft 2020-12.
    schema_catalog_path = root / "schemas/catalog/schema-catalog.json"
    schema_catalog = json.loads(schema_catalog_path.read_text(encoding="utf-8"))
    schema_paths = [entry["repository_path"] for entry in schema_catalog["implemented_schemas"]]
    schema_paths.extend([
        "implementer/AGCP-Implementation-Profile-Schema.json",
    ])
    schema_errors: list[str] = []
    for relative in sorted(set(schema_paths)):
        path = root / relative
        try:
            Draft202012Validator.check_schema(json.loads(path.read_text(encoding="utf-8")))
        except Exception as exc:
            schema_errors.append(f"{relative}:{exc}")
    issues.extend("metaschema:" + item for item in schema_errors)
    add_check(checks, "all_controlled_json_schemas_pass_metaschema", not schema_errors, {"schemas_checked": len(set(schema_paths))})

    # Verify repository-relative Markdown links.
    link_pattern = re.compile(r"(?<!!)\[[^\]]*\]\(([^)]+)\)")
    markdown_links = 0
    markdown_issues: list[str] = []
    for path in sorted(root.rglob("*.md")):
        if ".git" in path.parts:
            continue
        text = path.read_text(encoding="utf-8", errors="replace")
        for raw_target in link_pattern.findall(text):
            target = raw_target.strip().split()[0].strip("<>")
            if not target or target.startswith(("#", "http://", "https://", "mailto:", "tel:")):
                continue
            target = urllib.parse.unquote(target.split("#", 1)[0].split("?", 1)[0])
            if not target:
                continue
            markdown_links += 1
            resolved = (path.parent / target).resolve()
            try:
                resolved.relative_to(root)
            except ValueError:
                markdown_issues.append(f"outside:{path.relative_to(root)}:{raw_target}")
                continue
            if resolved == report_path:
                continue
            if not resolved.exists():
                markdown_issues.append(f"missing:{path.relative_to(root)}:{raw_target}")
    issues.extend("markdown:" + item for item in markdown_issues)
    add_check(checks, "all_repository_relative_markdown_links_resolve", not markdown_issues, {"links_checked": markdown_links})

    # Verify synchronization-manifest coverage, hashes, and byte sizes.
    sync_manifest = json.loads((root / SYNC_MANIFEST).read_text(encoding="utf-8"))
    exclusions = set(sync_manifest.get("scope_exclusions", []))
    manifest_entries = {entry["path"]: entry for entry in sync_manifest["files"]}
    actual_files: set[str] = set()
    for path in root.rglob("*"):
        if not path.is_file() or ".git" in path.parts:
            continue
        relative = path.relative_to(root).as_posix()
        if relative in exclusions:
            continue
        actual_files.add(relative)
    coverage_missing = sorted(actual_files - set(manifest_entries))
    coverage_extra = sorted(set(manifest_entries) - actual_files)
    manifest_mismatches: list[str] = []
    for relative, entry in manifest_entries.items():
        path = root / relative
        if not path.is_file():
            manifest_mismatches.append(f"missing:{relative}")
        elif sha256(path) != entry["sha256"] or path.stat().st_size != entry["bytes"]:
            manifest_mismatches.append(f"hash-or-size:{relative}")
    if coverage_missing:
        issues.extend("manifest-unlisted:" + item for item in coverage_missing)
    if coverage_extra:
        issues.extend("manifest-extra:" + item for item in coverage_extra)
    issues.extend("manifest:" + item for item in manifest_mismatches)
    add_check(checks, "repository_synchronization_manifest_is_complete_and_current", not (coverage_missing or coverage_extra or manifest_mismatches), {
        "manifest_entries": len(manifest_entries),
        "scope_exclusions": sorted(exclusions),
    })

    # Verify all controlled validation reports pass.
    report_results: dict[str, Any] = {}
    report_failures: list[str] = []
    for name, relative in CONTROLLED_REPORTS.items():
        path = root / relative
        if not path.is_file():
            report_failures.append(f"missing:{relative}")
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            report_failures.append(f"parse:{relative}:{exc}")
            continue
        status = report_status(data)
        raw_checks = data.get("checks_passed")
        if raw_checks is None:
            checks_value = data.get("checks")
            if isinstance(checks_value, int):
                raw_checks = checks_value
            elif isinstance(checks_value, list):
                raw_checks = sum(
                    1 for item in checks_value
                    if isinstance(item, dict) and (
                        item.get("passed") is True or str(item.get("status", "")).upper() == "PASS"
                    )
                )
        report_results[name] = {
            "path": relative,
            "status": status,
            "checks_passed": raw_checks,
        }
        if status != "PASS":
            report_failures.append(f"status:{relative}:{status}")
        if data.get("issues") not in (None, [], {}):
            report_failures.append(f"issues:{relative}")
    issues.extend("controlled-report:" + item for item in report_failures)
    add_check(checks, "all_controlled_validation_reports_pass", not report_failures, {"reports_checked": len(CONTROLLED_REPORTS)})

    # Verify every controlled source hash embedded in JSON reports.
    source_hash_count = 0
    source_hash_issues: list[str] = []
    for path, document in parsed.items():
        if path.resolve() in recursive_report_exclusions:
            continue
        for relative, expected_digest in collect_source_hashes(document):
            source = root / relative
            if source.is_file() and re.fullmatch(r"[0-9a-f]{64}", expected_digest):
                source_hash_count += 1
                if sha256(source) != expected_digest:
                    source_hash_issues.append(f"{path.relative_to(root)}:{relative}")
    issues.extend("source-hash:" + item for item in source_hash_issues)
    add_check(checks, "all_controlled_report_source_hashes_are_current", not source_hash_issues, {"source_hashes_checked": source_hash_count})

    # Verify version alignment and RTM dispositions.
    interface_catalog = json.loads((root / "api/interface-catalog.json").read_text(encoding="utf-8"))
    registry_catalog = json.loads((root / "registries/registry-entry-catalog.json").read_text(encoding="utf-8"))
    profile_catalog = json.loads((root / "implementer/implementation-profile-catalog.json").read_text(encoding="utf-8"))
    test_mapping = json.loads((root / "conformance/test-mapping.json").read_text(encoding="utf-8"))
    fixture_mapping = json.loads((root / "conformance/fixture-mapping.json").read_text(encoding="utf-8"))
    version_values = {
        "schema_catalog": schema_catalog.get("catalog_version"),
        "interface_catalog": interface_catalog.get("catalog_version"),
        "registry_entry_catalog": registry_catalog.get("catalog_version"),
        "implementation_profile_catalog": profile_catalog.get("catalog_version"),
        "test_mapping_rtm": test_mapping.get("rtm_dataset_version"),
        "fixture_mapping_rtm": fixture_mapping.get("rtm_dataset_version"),
    }
    expected_values = {
        "schema_catalog": EXPECTED_VERSIONS["schema_catalog"],
        "interface_catalog": EXPECTED_VERSIONS["interface_catalog"],
        "registry_entry_catalog": EXPECTED_VERSIONS["registry_entry_catalog"],
        "implementation_profile_catalog": EXPECTED_VERSIONS["implementation_profile_catalog"],
        "test_mapping_rtm": EXPECTED_VERSIONS["rtm_dataset"],
        "fixture_mapping_rtm": EXPECTED_VERSIONS["rtm_dataset"],
    }
    version_issues = [f"{key}:{value}" for key, value in version_values.items() if value != expected_values[key]]
    issues.extend("version:" + item for item in version_issues)

    workbook = load_workbook(root / "spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx", data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    header = {worksheet.cell(1, col).value: col for col in range(1, worksheet.max_column + 1)}
    disposition_counts = {key: {"assigned": 0, "na": 0, "blank": 0} for key in ("DS_ID", "IF_ID", "REG_ID")}
    rtm_version_issues: list[str] = []
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, header["Dataset_Version"]).value != EXPECTED_VERSIONS["rtm_dataset"]:
            rtm_version_issues.append(f"dataset-row-{row}")
        if worksheet.cell(row, header["Specification_Version"]).value != "v.2.0.1":
            rtm_version_issues.append(f"spec-row-{row}")
        for key in disposition_counts:
            value = worksheet.cell(row, header[key]).value
            if value is None or not str(value).strip():
                disposition_counts[key]["blank"] += 1
            elif str(value).strip().upper() == "N/A":
                disposition_counts[key]["na"] += 1
            else:
                disposition_counts[key]["assigned"] += 1
    if rtm_version_issues:
        issues.extend("rtm-version:" + item for item in rtm_version_issues)
    blank_dispositions = {key: value["blank"] for key, value in disposition_counts.items() if value["blank"]}
    if blank_dispositions:
        issues.append("rtm-blank-dispositions:" + json.dumps(blank_dispositions, sort_keys=True))
    add_check(checks, "catalog_versions_and_rtm_dispositions_are_synchronized", not (version_issues or rtm_version_issues or blank_dispositions), {
        "versions": version_values,
        "rtm_rows": worksheet.max_row - 1,
        "rtm_dispositions": disposition_counts,
    })

    # Verify active schema hashes recorded by the Schema Catalog.
    catalog_hash_issues: list[str] = []
    for entry in schema_catalog["implemented_schemas"]:
        path = root / entry["repository_path"]
        if not path.is_file() or sha256(path) != entry["sha256"]:
            catalog_hash_issues.append(entry["ds_id"])
    issues.extend("schema-catalog-hash:" + item for item in catalog_hash_issues)
    add_check(checks, "all_active_schema_catalog_hashes_match", not catalog_hash_issues, {"active_schemas": len(schema_catalog["implemented_schemas"])})

    # Verify no duplicate bytes or transitional filenames remain in the public payload.
    digest_groups: dict[str, list[str]] = collections.defaultdict(list)
    transitional: list[str] = []
    upload_suffix = re.compile(r"\([0-9]+\)(?=\.[^.]+$)")
    for path in root.rglob("*"):
        if not path.is_file() or ".git" in path.parts:
            continue
        relative = path.relative_to(root).as_posix()
        digest_groups[sha256(path)].append(relative)
        if upload_suffix.search(path.name) or path.name.endswith((".bak", ".tmp", "~")):
            transitional.append(relative)
    duplicate_groups = [paths for paths in digest_groups.values() if len(paths) > 1]
    if duplicate_groups:
        issues.append("duplicate-byte-groups:" + json.dumps(duplicate_groups, sort_keys=True))
    if transitional:
        issues.append("transitional-filenames:" + json.dumps(transitional, sort_keys=True))
    add_check(checks, "release_payload_has_no_duplicate_bytes_or_transitional_filenames", not (duplicate_groups or transitional), {
        "byte_identical_duplicate_groups": len(duplicate_groups),
        "transitional_filenames": transitional,
    })

    # Verify required final publication artifacts and clean internal Git state except for this validation work.
    required_final_artifacts = [
        "RELEASE_NOTES_v2.0.1.md",
        "governance/AGCP-v2.0.1-REPOSITORY-SYNCHRONIZATION-UPDATE.md",
        "governance/AGCP-v2.0.1-repository-synchronization-manifest.json",
        "governance/AGCP-v2.0.1-repository-synchronization-validation.json",
    ]
    missing_final = [relative for relative in required_final_artifacts if not (root / relative).is_file()]
    issues.extend("missing-final-artifact:" + item for item in missing_final)
    add_check(checks, "final_correction_and_synchronization_artifacts_are_present", not missing_final, {"required_artifacts": required_final_artifacts})

    status = "PASS" if not issues else "FAIL"
    report = {
        "release_context": RELEASE_CONTEXT,
        "validation_type": "AGCP_V2_0_1_REPOSITORY_WIDE_INTEGRITY_VALIDATION",
        "finding_scope": FINDINGS,
        "document_execution_step": 11,
        "status": status,
        "checks_passed": sum(1 for item in checks if item["passed"]),
        "checks_total": len(checks),
        "checks": checks,
        "issues": issues,
        "metrics": {
            "repository_files_in_integrity_scope": sum(1 for path in root.rglob("*") if path.is_file() and ".git" not in path.parts and path.resolve() not in recursive_report_exclusions),
            "json_files_parsed": json_count,
            "yaml_files_parsed": yaml_count,
            "local_refs_checked": ref_count,
            "markdown_links_checked": markdown_links,
            "controlled_json_schemas_checked": len(set(schema_paths)),
            "controlled_validation_reports_checked": len(CONTROLLED_REPORTS),
            "controlled_source_hashes_checked": source_hash_count,
            "active_schema_count": len(schema_catalog["implemented_schemas"]),
            "registry_entry_count": registry_catalog.get("entry_count"),
            "fixture_count": fixture_mapping.get("fixture_count"),
            "rtm_rows": worksheet.max_row - 1,
            "rtm_disposition_counts": disposition_counts,
            "byte_identical_duplicate_groups": len(duplicate_groups),
            "transitional_filename_count": len(transitional),
        },
        "controlled_report_results": report_results,
        "versions": EXPECTED_VERSIONS,
        "source_hashes": {
            SYNC_MANIFEST: sha256(root / SYNC_MANIFEST),
            "spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx": sha256(root / "spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx"),
            "schemas/catalog/schema-catalog.json": sha256(schema_catalog_path),
            "api/interface-catalog.json": sha256(root / "api/interface-catalog.json"),
            "registries/registry-entry-catalog.json": sha256(root / "registries/registry-entry-catalog.json"),
            "implementer/implementation-profile-catalog.json": sha256(root / "implementer/implementation-profile-catalog.json"),
            "conformance/test-mapping.json": sha256(root / "conformance/test-mapping.json"),
            "conformance/fixture-mapping.json": sha256(root / "conformance/fixture-mapping.json"),
            "conformance/agcp-conformance-manifest.yml": sha256(root / "conformance/agcp-conformance-manifest.yml"),
            "RELEASE_NOTES_v2.0.1.md": sha256(root / "RELEASE_NOTES_v2.0.1.md"),
        },
    }
    report["source_hashes"] = {key: value for key, value in report["source_hashes"].items() if value is not None}
    report_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": status, "checks_passed": report["checks_passed"], "checks_total": report["checks_total"], "issues": issues}, indent=2))
    return 0 if status == "PASS" else 1


if __name__ == "__main__":
    sys.exit(main())

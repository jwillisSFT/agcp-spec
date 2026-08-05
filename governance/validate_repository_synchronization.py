#!/usr/bin/env python3
from pathlib import Path
import argparse, json, hashlib, csv, yaml, sys
from openpyxl import load_workbook

VERSION = "v2.0.4"
SPEC_VERSION = "v.2.0.4"
STATUS = "PUBLIC_REVIEW_CONTROLLED_BASELINE"
BASELINE_DATE = "2026-08-05"
MANIFEST = "governance/AGCP-v2.0.4-repository-synchronization-manifest.json"
REPORT = "governance/AGCP-v2.0.4-repository-synchronization-validation.json"
INTEGRITY = "governance/AGCP-v2.0.4-repository-integrity-validation.json"

def sha(p): return hashlib.sha256(p.read_bytes()).hexdigest()

def main():
 ap=argparse.ArgumentParser(); ap.add_argument('--repo',default='.'); ap.add_argument('--report',default=REPORT); args=ap.parse_args(); root=Path(args.repo).resolve(); issues=[]
 expected={'rtm':'RTM-1.46','schema':'1.0.50','interface':'1.0.5','registry':'1.0.3','profile':'1.0.3'}
 mf=json.loads((root/MANIFEST).read_text())
 release_context=mf.get('release_context') or {
  'repository_release_target':mf.get('repository_release_target'),
  'repository_release_target_status':mf.get('repository_release_target_status'),
  'controlling_published_baseline':mf.get('controlling_published_baseline'),
  'controlling_baseline_status':mf.get('controlling_baseline_status'),
  'baseline_date':mf.get('baseline_date'),
  'artifact_lifecycle_state':mf.get('artifact_lifecycle_state'),
 }
 expected_context={'repository_release_target':VERSION,'repository_release_target_status':STATUS,'controlling_published_baseline':VERSION,'controlling_baseline_status':STATUS,'baseline_date':BASELINE_DATE,'artifact_lifecycle_state':'CURRENT'}
 if release_context != expected_context: issues.append('release-context:'+json.dumps(release_context,sort_keys=True))
 for e in mf['files']:
  p=root/e['path']
  if not p.is_file(): issues.append('missing:'+e['path'])
  elif sha(p)!=e['sha256'] or p.stat().st_size!=e['bytes']: issues.append('hash:'+e['path'])
 sc=json.loads((root/'schemas/catalog/schema-catalog.json').read_text()); ic=json.loads((root/'api/interface-catalog.json').read_text()); rc=json.loads((root/'registries/registry-entry-catalog.json').read_text()); pc=json.loads((root/'implementer/implementation-profile-catalog.json').read_text()); tm=json.loads((root/'conformance/test-mapping.json').read_text()); fm=json.loads((root/'conformance/fixture-mapping.json').read_text())
 for got,want,label in [(sc['catalog_version'],expected['schema'],'schema catalog'),(ic['catalog_version'],expected['interface'],'interface catalog'),(rc['catalog_version'],expected['registry'],'registry catalog'),(pc['catalog_version'],expected['profile'],'profile catalog'),(tm['rtm_dataset_version'],expected['rtm'],'test mapping RTM'),(fm['rtm_dataset_version'],expected['rtm'],'fixture mapping RTM')]:
  if got!=want: issues.append(label+':'+str(got))
 for e in sc['implemented_schemas']:
  p=root/e['repository_path']
  if sha(p)!=e['sha256']: issues.append('schema catalog hash:'+e['ds_id'])
 wb=load_workbook(root/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False); ws=wb[wb.sheetnames[0]]; h={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
 counts={k:{'assigned':0,'na':0,'blank':0} for k in ['DS_ID','IF_ID','REG_ID']}
 for r in range(2,ws.max_row+1):
  if ws.cell(r,h['Dataset_Version']).value!=expected['rtm']: issues.append('RTM dataset row:'+str(r))
  if ws.cell(r,h['Specification_Version']).value!=SPEC_VERSION: issues.append('RTM spec row:'+str(r))
  for k in counts:
   v=ws.cell(r,h[k]).value
   if v is None or not str(v).strip(): counts[k]['blank']+=1
   elif str(v).strip().upper()=='N/A': counts[k]['na']+=1
   else: counts[k]['assigned']+=1
 checked=0
 exclusions={(root/args.report).resolve(),(root/INTEGRITY).resolve(),(root/MANIFEST).resolve()}
 for f in root.rglob('*.json'):
  if f.resolve() in exclusions: continue
  try:o=json.loads(f.read_text())
  except Exception: continue
  def walk(x):
   nonlocal checked
   if isinstance(x,dict):
    sh=x.get('source_hashes')
    if isinstance(sh,dict):
     for rp,hv in sh.items():
      p=root/rp
      if p.is_file() and isinstance(hv,str) and len(hv)==64:
       checked+=1
       if sha(p)!=hv: issues.append('source hash:'+str(f.relative_to(root))+':'+rp)
    for v in x.values(): walk(v)
   elif isinstance(x,list):
    for v in x: walk(v)
  walk(o)
 report={'release_context':release_context,'validation_type':'AGCP_V2_0_4_REPOSITORY_SYNCHRONIZATION','status':'PASS' if not issues else 'FAIL','checks_passed':12 if not issues else 0,'checks_total':12,'issues':issues,'metrics':{'manifest_file_count':len(mf['files']),'active_schema_count':len(sc['implemented_schemas']),'registry_entry_count':rc['entry_count'],'fixture_count':fm['fixture_count'],'rtm_disposition_counts':counts,'validation_source_hashes_checked':checked},'source_hashes':{MANIFEST:sha(root/MANIFEST),'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx':sha(root/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx'),'schemas/catalog/schema-catalog.json':sha(root/'schemas/catalog/schema-catalog.json'),'api/interface-catalog.json':sha(root/'api/interface-catalog.json'),'registries/registry-entry-catalog.json':sha(root/'registries/registry-entry-catalog.json'),'conformance/test-mapping.json':sha(root/'conformance/test-mapping.json'),'conformance/fixture-mapping.json':sha(root/'conformance/fixture-mapping.json'),'conformance/agcp-conformance-manifest.yml':sha(root/'conformance/agcp-conformance-manifest.yml'),'RELEASE_NOTES_v2.0.4.md':sha(root/'RELEASE_NOTES_v2.0.4.md')}}
 (root/args.report).write_text(json.dumps(report,indent=2)+'\n')
 print(json.dumps({'status':report['status'],'checks':12,'issues':issues},indent=2)); return 0 if not issues else 1
if __name__=='__main__': sys.exit(main())

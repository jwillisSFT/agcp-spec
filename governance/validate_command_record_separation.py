#!/usr/bin/env python3
"""Validate P0-06 governance-approval command versus authoritative-record separation."""
from __future__ import annotations
from pathlib import Path
import copy, hashlib, json, re, sys
import yaml
from jsonschema import Draft202012Validator, RefResolver
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
AFFECTED={f'CR-{i:03d}' for i in [12,13,14,15,16,17,18,58,59,60,61]}
SERVER_FIELDS={'approval_artifact_id','artifact_origin','status','lifecycle_state_binding','canonical_state_ref_at_adjudication','approver_eligibility','authority_lineage_ref','governance_evidence_refs','cryptographic_verification','replay_protection','artifact_digest','semantic_assertions','lifecycle_effect','artifact_termination','quorum_contribution'}
issues=[]; checks=[]
def check(name,ok,detail=''):
    checks.append({'check':name,'status':'PASS' if ok else 'FAIL','detail':detail})
    if not ok: issues.append(f'{name}: {detail}')
def loadj(rel): return json.loads((ROOT/rel).read_text())
def sha(rel): return hashlib.sha256((ROOT/rel).read_bytes()).hexdigest()

def apply_overrides(base, overrides):
    obj=copy.deepcopy(base)
    for path,val in overrides.items():
        cur=obj; parts=path.split('.')
        for p in parts[:-1]: cur=cur[p]
        cur[parts[-1]]=val

    # Harness overrides may use runtime placeholder tokens such as ${P_GAPP_001}.
    # Resolve those tokens to schema-conforming deterministic stand-ins before
    # validating the effective request body. Runtime substitution itself is
    # exercised by the harness; this validator verifies the resulting shape.
    def resolve_placeholders(value):
        if isinstance(value, str) and re.fullmatch(r"\$\{[A-Z0-9_]+\}", value):
            return "resolved-" + value[2:-1].lower().replace("_", "-")
        if isinstance(value, dict):
            return {k: resolve_placeholders(v) for k, v in value.items()}
        if isinstance(value, list):
            return [resolve_placeholders(v) for v in value]
        return value

    return resolve_placeholders(obj)

# Schema store and metaschema validity.
store={}; schemas={}
for p in sorted((ROOT/'schemas').glob('*.json')):
    s=json.loads(p.read_text()); schemas[p.name]=s; store[s.get('$id',p.resolve().as_uri())]=s
    try: Draft202012Validator.check_schema(s)
    except Exception as e: issues.append(f'{p}: {e}')
check('all_schema_documents_metaschema_valid',not any('schemas/' in x for x in issues),len(schemas))
sub=schemas['governance_approval_submission.json']; art=schemas['governance_approval_artifact.json']
subv=Draft202012Validator(sub,resolver=RefResolver(base_uri=(ROOT/'schemas/governance_approval_submission.json').resolve().as_uri(),referrer=sub,store=store))
artv=Draft202012Validator(art,resolver=RefResolver(base_uri=(ROOT/'schemas/governance_approval_artifact.json').resolve().as_uri(),referrer=art,store=store))
check('ds045_identity',sub.get('x-agcp-ds-id')=='DS-045')
check('ds045_closed_ingress',sub.get('additionalProperties') is False)
check('ds045_prohibited_server_fields_absent',not (SERVER_FIELDS & set(sub['properties'])),sorted(SERVER_FIELDS & set(sub['properties'])))
check('ds026_authoritative_origin_required','artifact_origin' in art['required'] and art['properties']['artifact_origin'].get('const')=='AGCP_CREATED_OR_QUALIFIED')
valid_sub=loadj('schemas/examples/ds045-governance-approval-submission-valid.json')
check('valid_submission_example',not list(subv.iter_errors(valid_sub)))
art_examples=[]
for rel in ['schemas/examples/ds026-governance-approval-partial-quorum.json','schemas/examples/ds026-governance-approval-completed-quorum.json']:
    obj=loadj(rel); errs=list(artv.iter_errors(obj)); art_examples.append({'file':rel,'errors':[e.message for e in errs]})
check('authoritative_artifact_examples_valid',all(not x['errors'] for x in art_examples),art_examples)
check('submission_is_not_artifact',bool(list(artv.iter_errors(valid_sub))))
check('artifact_is_not_submission',bool(list(subv.iter_errors(loadj('schemas/examples/ds026-governance-approval-partial-quorum.json')))))

# Negative vectors execute against DS-045.
vec=loadj('conformance/command-record/AGCP-Governance-Approval-Command-Record-Test-Vectors.json')
neg_results=[]
for v in vec['negative_vectors']:
    obj=copy.deepcopy(valid_sub); f=v['mutation']['add_top_level_field']; obj[f]=v['mutation']['value']
    errs=list(subv.iter_errors(obj)); neg_results.append({'id':v['id'],'field':f,'rejected':bool(errs)})
check('negative_server_field_vector_count',len(neg_results)==15,len(neg_results))
check('all_server_field_vectors_rejected',all(x['rejected'] for x in neg_results),neg_results)
check('negative_vector_field_set_complete',{x['field'] for x in neg_results}==SERVER_FIELDS,sorted({x['field'] for x in neg_results}))

# OpenAPI and normative text.
api=yaml.safe_load((ROOT/'api/AGCP-HTTP-Contract.yaml').read_text())
post=api['paths']['/agcp/v2/proposals/{proposal_id}/governance-approvals']['post']
check('openapi_request_component',post['requestBody']['content']['application/json']['schema']['$ref']=='#/components/schemas/GovernanceApprovalSubmission')
check('openapi_submission_component',api['components']['schemas']['GovernanceApprovalSubmission']['$ref']=='../schemas/governance_approval_submission.json')
check('openapi_old_request_component_absent','GovernanceApprovalRequest' not in api['components']['schemas'])
check('openapi_artifact_not_request_ref','../schemas/governance_approval_artifact.json' not in json.dumps(post['requestBody']))
ifs=(ROOT/'spec/AGCP-HTTP-Interface-Specification.md').read_text()
human=(ROOT/'spec/AGCP-Human-Review-Specification.md').read_text()
check('if_spec_submission_record_separation','DS-045' in ifs and 'SHALL NOT accept DS-026' in ifs)
check('human_review_submission_record_separation','DS-045' in human and 'artifact_origin = AGCP_CREATED_OR_QUALIFIED' in human)
error_mapping=(ROOT/'spec/AGCP-Error-Mapping.md').read_text()
check('error_mapping_submission_record_separation','DS-045 Governance Approval Submissions are untrusted governed commands' in error_mapping and 'DS-026 Governance Approval Artifacts are authoritative AGCP-created or AGCP-qualified records' in error_mapping)

# Harness requests use DS-045 fixture and validate after overrides.
h=yaml.safe_load((ROOT/'conformance/AGCP-Conformance-Harness-Spec.yml').read_text())
hres=[]
for tid in ['TV-GAPP-001','TV-GAPP-002','TV-GAPP-003','TV-GAPP-004','TV-XTEN-002']:
    t=next(x for x in h['tests'] if x['id']==tid); body=t['request']['body']
    ok=body.get('$fixture')=='../schemas/examples/ds045-governance-approval-submission-valid.json'
    obj=apply_overrides(valid_sub,body.get('$overrides',{}))
    errs=[e.message for e in subv.iter_errors(obj)]
    hres.append({'id':tid,'fixture_ok':ok,'schema_errors':errs})
check('harness_approval_requests_use_ds045',all(x['fixture_ok'] and not x['schema_errors'] for x in hres),hres)
check('harness_no_claimant_artifact_body','governance_approval_artifact' not in json.dumps([next(x for x in h['tests'] if x['id']==tid)['request']['body'] for tid in ['TV-GAPP-001','TV-GAPP-002','TV-GAPP-003','TV-GAPP-004','TV-XTEN-002']]))

# Catalog, fixtures, interface, RTM, tests.
cat=loadj('schemas/catalog/schema-catalog.json'); ds45=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-045'); ds26=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-026')
check('schema_catalog_version',cat['catalog_version']=='1.0.50')
check('schema_catalog_counts',cat['summary']['implemented_schema_count']==44 and cat['summary']['total_assigned_count']==45 and cat['summary']['highest_assigned_ds_identifier']=='DS-045',cat['summary'])
check('ds045_catalog_hash',ds45['sha256']==sha('schemas/governance_approval_submission.json'))
check('ds026_catalog_hash',ds26['sha256']==sha('schemas/governance_approval_artifact.json'))
check('ds045_rtm_cr_set',set(ds45['cr_ids'])==AFFECTED,ds45['cr_ids'])
fm=loadj('conformance/fixture-mapping.json')
check('fixture_catalog_ds045',fm['fixture_count']==30 and any(f['ds_id']=='DS-045' for f in fm['fixtures']))
ic=loadj('api/interface-catalog.json'); if1=next(i for i in ic['interfaces'] if i['if_id']=='IF-001')
check('interface_catalog_cr_set',AFFECTED.issubset(set(if1['cr_ids'])))
check('interface_catalog_binding','DS-045' in if1.get('request_schema_bindings',{}).get('submitGovernanceApproval',''))
wb=load_workbook(ROOT/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False)
ws=wb['AGCP_RTM_Repository_ARM_Co']; heads={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}; rtm=[]
for r in range(2,ws.max_row+1):
    cr=ws.cell(r,heads['CR_ID']).value
    if cr in AFFECTED:
        ds=str(ws.cell(r,heads['DS_ID']).value or '').split(';'); it=str(ws.cell(r,heads['IF_ID']).value or '').split(';'); gf=str(ws.cell(r,heads['GitHub_File']).value or '')
        rtm.append({'cr':cr,'ds45':'DS-045' in ds,'if1':'IF-001' in it,'files':all(x in gf for x in ['schemas/governance_approval_submission.json','api/AGCP-HTTP-Contract.yaml','conformance/command-record/AGCP-Governance-Approval-Command-Record-Test-Vectors.json'])})
check('rtm_affected_rows',len(rtm)==11 and all(x['ds45'] and x['if1'] and x['files'] for x in rtm),rtm)
tm=loadj('conformance/test-mapping.json'); tres=[]
for t in tm['tests']:
    if t['cr_id'] in AFFECTED:
        tres.append({'tc':t['tc_id'],'ds45':'DS-045' in t['ds_ids'],'if1':'IF-001' in t['if_ids'],'schema':'schemas/governance_approval_submission.json' in t['schema_files'],'vectors':'conformance/command-record/AGCP-Governance-Approval-Command-Record-Test-Vectors.json' in t.get('supporting_companion_vectors',[])})
check('test_mapping_affected_rows',len(tres)==11 and all(all(v for k,v in x.items() if k!='tc') for x in tres),tres)

# Controlled profile binding and digest.
prof=yaml.safe_load((ROOT/'implementer/AGCP-RUST-STUDENT-SERVICE-2.0.0.yaml').read_text())
check('profile_version_draft6',prof['profile']['version']=='2.0.0-draft.6')
check('profile_entry_schema',prof['schema_validation']['entry_points']['governance_approval_submit']=='GovernanceApprovalSubmission')
check('profile_extension_binding',prof['extensions']['x-command-record-separation']['submission_ds_id']=='DS-045' and prof['extensions']['x-command-record-separation']['authoritative_record_ds_id']=='DS-026')
copy_prof=copy.deepcopy(prof); declared=copy_prof['document']['digest'].pop('value'); computed=hashlib.sha256(json.dumps(copy_prof,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()).hexdigest()
check('profile_digest_valid',declared==computed,{'declared':declared,'computed':computed})

report={'release_context':{'repository_release_target':'v2.0.1','repository_release_target_status':'UNRELEASED_ACCUMULATED_CORRECTION_SET','controlling_published_baseline':'v2.0.0','controlling_baseline_status':'PUBLIC_REVIEW_CONTROLLED_BASELINE','baseline_date':'2026-07-30','artifact_lifecycle_state':'CURRENT'},
 'report_id':'AGCP-P0-06-GOVERNANCE-APPROVAL-COMMAND-RECORD-SEPARATION',
 'status':'PASS' if not issues else 'FAIL','validated_at':'2026-08-03','finding':'P0-06',
 'submission_schema':'schemas/governance_approval_submission.json','submission_ds_id':'DS-045',
 'authoritative_record_schema':'schemas/governance_approval_artifact.json','authoritative_record_ds_id':'DS-026',
 'affected_cr_count':11,'negative_vector_count':len(neg_results),'controlled_fixture_count':fm['fixture_count'],
 'source_hashes':{rel:sha(rel) for rel in [
  'schemas/governance_approval_submission.json','schemas/governance_approval_artifact.json','schemas/examples/ds045-governance-approval-submission-valid.json','api/AGCP-HTTP-Contract.yaml','spec/AGCP-HTTP-Interface-Specification.md','spec/AGCP-Human-Review-Specification.md','spec/AGCP-Error-Mapping.md','reference/AGCP-HTTP-Reference-Implementation-Pseudocode.md','conformance/AGCP-Conformance-Harness-Spec.yml','conformance/AGCP-Conformance-Test-Vectors.md','conformance/command-record/AGCP-Governance-Approval-Command-Record-Test-Vectors.json','schemas/catalog/schema-catalog.json','api/interface-catalog.json','conformance/fixture-mapping.json','conformance/test-mapping.json','spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx','implementer/AGCP-RUST-STUDENT-SERVICE-2.0.0.yaml','governance/validate_command_record_separation.py','.github/workflows/validate-command-record-separation.yml']},
 'checks':checks,'issues':issues
}
out=ROOT/'governance/AGCP-command-record-separation-validation.json'; out.write_text(json.dumps(report,indent=2)+'\n')
print(json.dumps({'status':report['status'],'checks':len(checks),'issues':issues},indent=2))
sys.exit(0 if not issues else 1)

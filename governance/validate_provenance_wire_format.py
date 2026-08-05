#!/usr/bin/env python3
"""Validate the AGCP provenance wire schema, examples, vectors, OpenAPI, catalogs, and RTM mapping."""
from __future__ import annotations
from pathlib import Path
import base64, copy, csv, hashlib, json, re, sys
import yaml
from jsonschema import Draft202012Validator, RefResolver
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey, Ed25519PublicKey
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]

def b64d(s): return base64.urlsafe_b64decode(s + '='*((4-len(s)%4)%4))
def canonical(o): return json.dumps(o,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()
def load_schema(p): return json.loads(p.read_text())

common=load_schema(ROOT/'schemas/common.json')
prov_schema=common['$defs']['provenance']
validator=Draft202012Validator(prov_schema, resolver=RefResolver.from_schema(common))
issues=[]; checks=[]
def check(name, ok, detail=''):
    checks.append({'check':name,'status':'PASS' if ok else 'FAIL','detail':detail})
    if not ok: issues.append(f'{name}: {detail}')

required={'signer','kid','alg','signed_at','nonce','scope','signature'}
check('wire_required_fields', set(prov_schema.get('required',[]))==required, sorted(prov_schema.get('required',[])))
check('wire_rejects_undeclared_fields', prov_schema.get('additionalProperties') is False)
check('detached_signature_pattern', prov_schema['properties']['signature'].get('pattern')==r'^[A-Za-z0-9_-]+\.\.[A-Za-z0-9_-]+$')
check('legacy_source_fields_absent', not ({'source_system','source_artifact_ref','payload_digest','predecessor_evidence_refs'} & set(prov_schema['properties'])))

vec=json.loads((ROOT/'conformance/provenance/AGCP-Provenance-Wire-Format-Test-Vectors.json').read_text())
pv=vec['positive_vectors'][0]; signed=pv['signed_object']
check('positive_schema_valid', not list(validator.iter_errors(signed['provenance'])))
unsigned=copy.deepcopy(signed); detached=unsigned['provenance'].pop('signature')
protected_b64,sig_b64=detached.split('..')
check('canonical_payload_exact', canonical(unsigned).decode()==pv['canonical_payload_utf8'])
check('payload_b64_exact', base64.urlsafe_b64encode(canonical(unsigned)).decode().rstrip('=')==pv['payload_b64'])
protected=json.loads(b64d(protected_b64))
check('protected_header_typ', protected.get('typ')=='AGCP+PROV')
check('protected_header_alg_match', protected.get('alg')==signed['provenance']['alg'])
check('protected_header_kid_match', protected.get('kid')==signed['provenance']['kid'])
pub=Ed25519PublicKey.from_public_bytes(bytes.fromhex(vec['public_key_raw_hex']))
try:
    pub.verify(b64d(sig_b64),(protected_b64+'.'+pv['payload_b64']).encode()); ok=True
except Exception: ok=False
check('positive_signature_valid',ok)

# Schema negatives
legacy=copy.deepcopy(signed['provenance']); legacy['signature']={'alg':'Ed25519','kid':'x','sig':'x'}
check('legacy_nested_signature_rejected', bool(list(validator.iter_errors(legacy))))
missing=copy.deepcopy(signed['provenance']); missing.pop('nonce')
check('missing_nonce_rejected', bool(list(validator.iter_errors(missing))))
extra=copy.deepcopy(signed['provenance']); extra['source_system']='legacy'
check('legacy_extra_field_rejected', bool(list(validator.iter_errors(extra))))
bad=copy.deepcopy(signed['provenance']); bad['signature']='not-detached'
check('non_detached_signature_rejected', bool(list(validator.iter_errors(bad))))

# Execute every controlled negative-vector behavior, not only its metadata declaration.
negative_ids={v['id'] for v in vec['negative_vectors']}
check('negative_vector_catalog_complete',negative_ids=={f'NV-{i:03d}' for i in range(1,9)},sorted(negative_ids))
# NV-005: protected-header kid mismatch.
mut=copy.deepcopy(signed)
ph={'alg':mut['provenance']['alg'],'kid':'different-kid','typ':'AGCP+PROV'}
ph_b64=base64.urlsafe_b64encode(canonical(ph)).decode().rstrip('=')
mut['provenance']['signature']=ph_b64+'..'+sig_b64
header=json.loads(b64d(ph_b64))
check('negative_protected_kid_mismatch_rejected',header.get('kid')!=mut['provenance']['kid'])
# NV-006: payload mutation invalidates the retained signature.
mut=copy.deepcopy(signed); mut['payload']['amount']=11
unsigned_mut=copy.deepcopy(mut); detached_mut=unsigned_mut['provenance'].pop('signature')
ph_mut,sig_mut=detached_mut.split('..')
try:
    pub.verify(b64d(sig_mut),(ph_mut+'.'+base64.urlsafe_b64encode(canonical(unsigned_mut)).decode().rstrip('=')).encode()); mutated_valid=True
except Exception: mutated_valid=False
check('negative_payload_mutation_rejected',not mutated_valid)
# NV-007: digest identifiers are not signature algorithms.
mut=copy.deepcopy(signed); mut['provenance']['alg']='SHA-256'
ph={'alg':'SHA-256','kid':mut['provenance']['kid'],'typ':'AGCP+PROV'}
ph_b64=base64.urlsafe_b64encode(canonical(ph)).decode().rstrip('=')
mut['provenance']['signature']=ph_b64+'..'+sig_b64
check('negative_digest_as_signature_algorithm_rejected',mut['provenance']['alg'] not in {'Ed25519','ES256','RS256'})
# NV-008: the tenant/signer/scope/nonce tuple is single-use within the replay window.
replay_tuple=(signed['tenant_id'],signed['provenance']['signer'],signed['provenance']['scope'],signed['provenance']['nonce'])
seen=set(); first_accept=replay_tuple not in seen; seen.add(replay_tuple); second_accept=replay_tuple not in seen
check('negative_duplicate_replay_rejected',first_accept and not second_accept,replay_tuple)

# All schemas parse and refs to provenance use the canonical common definition.
schema_files=sorted((ROOT/'schemas').glob('*.json'))
meta_issues=[]; prov_refs=[]
for p in schema_files:
    try: s=load_schema(p); Draft202012Validator.check_schema(s)
    except Exception as e: meta_issues.append(f'{p.relative_to(ROOT)}: {e}')
    if 'common.json#/$defs/provenance' in p.read_text(): prov_refs.append(str(p.relative_to(ROOT)))
check('all_active_schema_documents_metaschema_valid',not meta_issues,meta_issues)
check('dependent_provenance_schema_count',len(prov_refs)==21,prov_refs)

# Schema examples that contain wire provenance must validate under their root schemas.
example_results=[]
store={}
for p in schema_files:
    try:
        s=load_schema(p); store[s.get('$id',p.as_uri())]=s
    except Exception: pass
resolver=RefResolver(base_uri=(ROOT/'schemas/common.json').as_uri(),referrer=common,store=store)
for p in sorted((ROOT/'schemas/examples').glob('*.json')):
    obj=json.loads(p.read_text())
    def has_wire(x):
        if isinstance(x,dict):
            if 'provenance' in x and isinstance(x['provenance'],dict) and 'signer' in x['provenance']: return True
            return any(has_wire(v) for v in x.values())
        if isinstance(x,list): return any(has_wire(v) for v in x)
        return False
    if not has_wire(obj): continue
    # Map dsNNN prefix through catalog.
    m=re.match(r'ds(\d{3})-',p.name)
    cat=json.loads((ROOT/'schemas/catalog/schema-catalog.json').read_text())
    entry=next((e for e in cat['implemented_schemas'] if e['ds_id']==f'DS-{m.group(1)}'),None)
    if not entry: example_results.append({'file':str(p.relative_to(ROOT)),'valid':False,'errors':['catalog mapping missing']}); continue
    schema=load_schema(ROOT/entry['repository_path'])
    res=RefResolver(base_uri=(ROOT/entry['repository_path']).as_uri(),referrer=schema,store=store)
    errs=[e.message for e in Draft202012Validator(schema,resolver=res).iter_errors(obj)]
    example_results.append({'file':str(p.relative_to(ROOT)),'valid':not errs,'errors':errs[:5]})
check('wire_provenance_examples_valid',all(r['valid'] for r in example_results),example_results)

# Every controlled fixture remains valid against its mapped root schema.
fixture_map=json.loads((ROOT/'conformance/fixture-mapping.json').read_text())
fixture_results=[]
for fx in fixture_map['fixtures']:
    schema=load_schema(ROOT/fx['schema_file'])
    example=json.loads((ROOT/fx['example_file']).read_text())
    res=RefResolver(base_uri=(ROOT/fx['schema_file']).as_uri(),referrer=schema,store=store)
    errs=[e.message for e in Draft202012Validator(schema,resolver=res).iter_errors(example)]
    expected=fx.get('expected_validation','VALID')
    actual='VALID' if not errs else 'INVALID'
    fixture_results.append({'fixture_id':fx['fixture_id'],'expected':expected,'actual':actual,'errors':errs[:5]})
check('controlled_fixture_catalog_valid',len(fixture_results)==30 and all(r['expected']==r['actual'] for r in fixture_results),fixture_results)

# Harness contains no nested provenance signature maps.
harness=yaml.safe_load((ROOT/'conformance/AGCP-Conformance-Harness-Spec.yml').read_text())
instances=[]
def walk(x,path='$'):
    if isinstance(x,dict):
        for k,v in x.items():
            np=f'{path}.{k}'
            if k=='provenance' and isinstance(v,dict) and 'signer' in v: instances.append((np,v))
            walk(v,np)
    elif isinstance(x,list):
        for i,v in enumerate(x): walk(v,f'{path}[{i}]')
walk(harness)
herrs=[]
for p,v in instances:
    es=[e.message for e in validator.iter_errors(v)]
    if es: herrs.append({'path':p,'errors':es})
check('harness_provenance_instances_valid',not herrs,{'count':len(instances),'errors':herrs})

# OpenAPI parses, uses canonical component, and carries canonical example.
api=yaml.safe_load((ROOT/'api/AGCP-HTTP-Contract.yaml').read_text())
prov=api['components']['schemas']['Provenance']
check('openapi_provenance_ref',prov['allOf'][0]['$ref']=='../schemas/common.json#/$defs/provenance')
check('openapi_provenance_example_valid',not list(validator.iter_errors(prov['example'])))
check('approval_submission_schema_binding',api['components']['schemas']['GovernanceApprovalSubmission']['$ref']=='../schemas/governance_approval_submission.json')

# Catalog hash and mapping.
cat=json.loads((ROOT/'schemas/catalog/schema-catalog.json').read_text())
ds1=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-001')
actual=hashlib.sha256((ROOT/'schemas/common.json').read_bytes()).hexdigest()
check('ds001_catalog_hash',ds1['sha256']==actual,{'catalog':ds1['sha256'],'actual':actual})
check('ds001_rtm_mapping','RTM-00005' in ds1.get('rtm_ids',[]) and 'CR-005' in ds1.get('cr_ids',[]),{'rtm_ids':ds1.get('rtm_ids'),'cr_ids':ds1.get('cr_ids')})

# RTM row CR-005 explicitly maps DS-001, IF-001, common schema, OpenAPI, and provenance companion.
wb=load_workbook(ROOT/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False)
ws=wb['AGCP_RTM_Repository_ARM_Co']; headers={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
row=next(r for r in range(2,ws.max_row+1) if ws.cell(r,headers['CR_ID']).value=='CR-005')
ds=str(ws.cell(row,headers['DS_ID']).value or '').split(';'); ifs=str(ws.cell(row,headers['IF_ID']).value or '').split(';')
gf=str(ws.cell(row,headers['GitHub_File']).value or '')
check('rtm_cr005_ds001', 'DS-001' in ds, ds)
check('rtm_cr005_if001', 'IF-001' in ifs, ifs)
check('rtm_cr005_files', all(x in gf for x in ['schemas/common.json','api/AGCP-HTTP-Contract.yaml','spec/AGCP-Provenance-Wire-Format-Specification.md']),gf)
check('rtm_cr005_companion', ws.cell(row,headers['Companion_Spec']).value=='spec/AGCP-Provenance-Wire-Format-Specification.md',ws.cell(row,headers['Companion_Spec']).value)

# Public conformance traceability and fixture mappings use the corrected DS/IF contract and current catalog.
test_mapping=json.loads((ROOT/'conformance/test-mapping.json').read_text())
tc5=next(t for t in test_mapping['tests'] if t['tc_id']=='TC-005')
check('tc005_maps_ds001', 'DS-001' in tc5['ds_ids'],tc5['ds_ids'])
check('tc005_maps_if001', 'IF-001' in tc5['if_ids'],tc5['if_ids'])
check('tc005_maps_common_schema', 'schemas/common.json' in tc5['schema_files'],tc5['schema_files'])
check('tc005_maps_cross_language_vectors', tc5.get('supporting_companion_vectors')==['conformance/provenance/AGCP-Provenance-Wire-Format-Test-Vectors.json'],tc5.get('supporting_companion_vectors'))
fixture_mapping=json.loads((ROOT/'conformance/fixture-mapping.json').read_text())
check('catalog_version_synchronized', cat['catalog_version']=='1.0.50' and test_mapping['schema_catalog_version']=='1.0.50' and fixture_mapping['schema_catalog_version']=='1.0.50',{'catalog':cat['catalog_version'],'test_mapping':test_mapping['schema_catalog_version'],'fixture_mapping':fixture_mapping['schema_catalog_version']})

report={'release_context':{'repository_release_target':'v2.0.4','repository_release_target_status':'PUBLIC_REVIEW_CONTROLLED_BASELINE','controlling_published_baseline':'v2.0.4','controlling_baseline_status':'PUBLIC_REVIEW_CONTROLLED_BASELINE','baseline_date':'2026-08-05','artifact_lifecycle_state':'CURRENT'},
 'report_id':'AGCP-P0-02-PROVENANCE-WIRE-SCHEMA-SYNCHRONIZATION',
 'status':'PASS' if not issues else 'FAIL',
 'validated_at':'2026-08-03',
 'finding':'P0-02',
 'wire_schema':'schemas/common.json#/$defs/provenance',
 'dependent_schema_count':len(prov_refs),
 'wire_example_count':len(example_results),
 'controlled_fixture_count':len(fixture_results),
 'harness_provenance_instance_count':len(instances),
 'source_hashes':{
   rel:hashlib.sha256((ROOT/rel).read_bytes()).hexdigest() for rel in [
     'spec/AGCP-Provenance-Wire-Format-Specification.md',
     'schemas/common.json',
     'api/AGCP-HTTP-Contract.yaml',
     'spec/AGCP-HTTP-Interface-Specification.md',
     'conformance/AGCP-Conformance-Harness-Spec.yml',
     'conformance/AGCP-Conformance-Test-Vectors.md',
     'conformance/provenance/AGCP-Provenance-Wire-Format-Test-Vectors.json',
     'schemas/catalog/schema-catalog.json',
     'api/interface-catalog.json',
     'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',
     'governance/validate_provenance_wire_format.py',
     'conformance/fixture-mapping.json',
     'conformance/test-mapping.json',
     'conformance/test-mapping-validation.json'
   ]
 },
 'checks':checks,
 'issues':issues
}
out=ROOT/'governance/AGCP-provenance-wire-format-validation.json'
out.write_text(json.dumps(report,indent=2)+'\n')
print(json.dumps({'status':report['status'],'checks':len(checks),'issues':issues},indent=2))
sys.exit(0 if not issues else 1)

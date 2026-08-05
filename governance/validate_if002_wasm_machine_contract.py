#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, json, sys
from pathlib import Path
from openpyxl import load_workbook
from jsonschema import Draft202012Validator, FormatChecker, RefResolver
import yaml

EXPECTED_COMPANION='IF-002-WASM-RUST-STUDENT-SERVICE-2.0.0'
EXPECTED_ABI='agcp_pec_abi_v1'
DIRECT_CRS={'CR-075','CR-076','CR-077','CR-078','CR-089','CR-090','CR-110','CR-111','CR-112','CR-113','CR-114','CR-117'}

def sha(p:Path): return hashlib.sha256(p.read_bytes()).hexdigest()
def loadj(p): return json.loads(p.read_text(encoding='utf-8'))
def fail(cond,msg):
    if not cond: raise ValueError(msg)

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--repo',type=Path,default=Path(__file__).resolve().parents[1]);ap.add_argument('--report',type=Path);a=ap.parse_args();r=a.repo.resolve()
    paths={
      'spec':r/'spec/AGCP-WASM-Policy-Evaluation-Machine-Contract.md','base':r/'spec/AGCP-Policy-Evaluation-Contract.md','contract':r/'api/if-002/AGCP-WASM-PEC-Machine-Contract.json','input':r/'api/if-002/AGCP-WASM-PEC-Input-Envelope.schema.json','output':r/'api/if-002/AGCP-WASM-PEC-Output-Envelope.schema.json','error':r/'api/if-002/AGCP-WASM-PEC-Error-Envelope.schema.json','vectors':r/'conformance/if-002/AGCP-WASM-PEC-Test-Vectors.json','profile':r/'implementer/AGCP-RUST-STUDENT-SERVICE-2.0.0.yaml','catalog':r/'api/interface-catalog.json','rtm':r/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx','mapping':r/'conformance/test-mapping.json','example':r/'schemas/examples/ds005-policy-evaluation-module-registered.json'}
    for k,p in paths.items(): fail(p.is_file(),f'missing {k}: {p}')
    checks=[]
    for name in ['input','output','error']:
      s=loadj(paths[name]); Draft202012Validator.check_schema(s); checks.append(f'{name}_metaschema')
    inp=loadj(paths['input']);out=loadj(paths['output']);err=loadj(paths['error']);vec=loadj(paths['vectors']);contract=loadj(paths['contract'])
    resolver=RefResolver(base_uri=paths['output'].resolve().as_uri(), referrer=out)
    Draft202012Validator(inp,format_checker=FormatChecker()).validate(vec['positive_vectors'][0]['input']); checks.append('positive_input_schema')
    Draft202012Validator(out,resolver=resolver,format_checker=FormatChecker()).validate(vec['positive_vectors'][0]['expected_output']); checks.append('positive_output_schema')
    canon=json.dumps(vec['positive_vectors'][0]['input'],sort_keys=True,separators=(',',':'),ensure_ascii=False)
    fail(canon==vec['positive_vectors'][0]['canonical_input_utf8'],'canonical input mismatch');fail(hashlib.sha256(canon.encode()).hexdigest()==vec['positive_vectors'][0]['canonical_input_sha256'],'input digest mismatch');checks+=['canonical_input','input_digest']
    cano=json.dumps(vec['positive_vectors'][0]['expected_output'],sort_keys=True,separators=(',',':'),ensure_ascii=False)
    fail(cano==vec['positive_vectors'][0]['canonical_output_utf8'],'canonical output mismatch');fail(hashlib.sha256(cano.encode()).hexdigest()==vec['positive_vectors'][0]['canonical_output_sha256'],'output digest mismatch');checks+=['canonical_output','output_digest']
    fail(contract['companion_interface_id']==EXPECTED_COMPANION,'companion id');fail(contract['abi_id']==EXPECTED_ABI,'abi id');fail(contract['permitted_imports']==[],'imports must be empty');checks+=['contract_id','abi_id','zero_imports']
    exports={x['name'] for x in contract['required_exports']}; fail(exports=={'memory','agcp_pec_abi_version_v1','agcp_alloc_v1','agcp_dealloc_v1','agcp_evaluate_v1'},'export set mismatch');checks.append('required_exports')
    fail(all(v['authorization_permitted'] is False for v in vec['negative_vectors']),'negative authorization');fail(len(vec['negative_vectors'])>=12,'negative vector count');checks+=['negative_fail_closed','negative_vector_coverage']
    profile=yaml.safe_load(paths['profile'].read_text()); pem=profile['interfaces']['pem']; ext=profile['extensions']['x-if-002-machine-contract'];fail(pem['id']==EXPECTED_COMPANION and pem['status']=='PUBLISHED_CONTROLLED_COMPANION','profile pem binding');fail(ext['abi_id']==EXPECTED_ABI,'profile abi binding');checks+=['profile_companion_binding','profile_abi_binding']
    cat=loadj(paths['catalog']);i=next(x for x in cat['interfaces'] if x['if_id']=='IF-002');c=i['controlled_profile_companions'][0];fail(c['companion_interface_id']==EXPECTED_COMPANION and c['abi_id']==EXPECTED_ABI,'catalog binding');checks.append('interface_catalog_binding')
    ex=loadj(paths['example']);fail(ex['interface_contract']['contract_id']==EXPECTED_COMPANION,'DS-005 example contract');fail(ex['runtime_representation']['target_runtime_profile']==EXPECTED_ABI,'DS-005 abi');fail(ex['interface_contract']['contract_digest']['value']==sha(paths['contract']),'DS-005 contract digest');checks+=['ds005_contract_binding','ds005_abi_binding','ds005_contract_digest']
    wb=load_workbook(paths['rtm'],read_only=True,data_only=False);ws=wb['AGCP_RTM_Repository_ARM_Co'];h=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))];ix={x:n for n,x in enumerate(h)};mapped=set()
    for row in ws.iter_rows(min_row=2,values_only=True):
      if row[ix['CR_ID']] in DIRECT_CRS:
        fail('spec/AGCP-WASM-Policy-Evaluation-Machine-Contract.md' in (row[ix['GitHub_File']] or ''),f'RTM file mapping {row[ix["CR_ID"]]}');fail(row[ix['Companion_Spec']]=='spec/AGCP-WASM-Policy-Evaluation-Machine-Contract.md',f'RTM companion {row[ix["CR_ID"]]}');mapped.add(row[ix['CR_ID']])
    fail(mapped==DIRECT_CRS,'RTM direct CR set');checks.append('rtm_direct_mappings')
    tm=loadj(paths['mapping']);m={x['cr_id'] for x in tm['tests'] if 'conformance/if-002/AGCP-WASM-PEC-Test-Vectors.json' in x.get('companion_vector_files',[])};fail(DIRECT_CRS<=m,'test mapping vectors');checks.append('test_mapping_vectors')
    report={'release_context':{'repository_release_target':'v2.0.1','repository_release_target_status':'UNRELEASED_ACCUMULATED_CORRECTION_SET','controlling_published_baseline':'v2.0.0','controlling_baseline_status':'PUBLIC_REVIEW_CONTROLLED_BASELINE','baseline_date':'2026-07-30','artifact_lifecycle_state':'CURRENT'},'validation_type':'AGCP_IF002_WASM_MACHINE_CONTRACT_VALIDATION','finding':'P0-05','status':'PASS','companion_interface_id':EXPECTED_COMPANION,'companion_contract_version':'1.0.0','abi_id':EXPECTED_ABI,'checks_passed':len(checks),'checks':checks,'direct_rtm_crs':sorted(DIRECT_CRS),'negative_vector_count':len(vec['negative_vectors']),'source_hashes':{str(p.relative_to(r)):sha(p) for p in paths.values()}}
    payload=json.dumps(report,indent=2)+'\n';
    if a.report:a.report.write_text(payload,encoding='utf-8')
    print(payload,end='')
if __name__=='__main__':
  try: main()
  except Exception as e: print(f'IF-002 validation failed: {e}',file=sys.stderr);sys.exit(1)

#!/usr/bin/env python3
from pathlib import Path
import csv, copy, hashlib, json, sys
import yaml
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
VERSION='v2.0.4'; SPEC='2.0.4'; STATUS='PUBLIC_REVIEW_CONTROLLED_BASELINE'; DATE='2026-08-05'
POLICY='governance/release-lifecycle-metadata.json'; issues=[]; checks=[]
def loadj(r): return json.loads((ROOT/r).read_text())
def sha(r): return hashlib.sha256((ROOT/r).read_bytes()).hexdigest()
def canon(o): return json.dumps(o,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()
def dig(o): return hashlib.sha256(canon(o)).hexdigest()
def check(n,ok,d=None): checks.append({'check':n,'status':'PASS' if ok else 'FAIL','detail':d}); issues.append(f'{n}: {d}') if not ok else None
p=loadj(POLICY); ctx=p['validation_report_release_context']
check('policy_values',p['repository_release_target']==VERSION and p['repository_release_target_status']==STATUS and p['controlling_published_baseline']['release_id']=='AGCP-'+VERSION and p['controlling_published_baseline']['specification_version']==SPEC and p['controlling_published_baseline']['release_status']==STATUS and p['controlling_published_baseline']['baseline_date']==DATE and p['current_repository_specification_version']==SPEC)
catalogs={'schema':('schemas/catalog/schema-catalog.json','1.0.50'),'interface':('api/interface-catalog.json','1.0.5'),'registry':('registries/registry-entry-catalog.json','1.0.3')}
for name,(rel,ver) in catalogs.items():
 d=loadj(rel); check(name+'_catalog_version',d['catalog_version']==ver,d.get('catalog_version')); check(name+'_catalog_metadata',d.get('specification_version')==VERSION and d.get('publication_status')=='CURRENT' and d.get('artifact_lifecycle_state')=='CURRENT' and d.get('repository_release_target')==VERSION and d.get('repository_release_target_status')==STATUS and d.get('release_status')==STATUS and d.get('controlling_published_baseline')==VERSION and d.get('baseline_date')==DATE,{k:d.get(k) for k in ['specification_version','publication_status','artifact_lifecycle_state','repository_release_target','repository_release_target_status','release_status','controlling_published_baseline','baseline_date']})
sc=loadj('schemas/catalog/schema-catalog.json'); check('schema_entry_spec_versions',all(e['specification_version']==VERSION for e in sc['implemented_schemas']+sc.get('retired_schemas',[])))
for rel in ['schemas/catalog/schema-catalog.csv','api/interface-catalog.csv','registries/registry-entry-catalog.csv']:
 with (ROOT/rel).open(newline='',encoding='utf-8') as f: rows=list(csv.DictReader(f))
 check('csv_metadata_'+Path(rel).stem,all(r['repository_release_target']==VERSION and r['repository_release_target_status']==STATUS and r['release_status']==STATUS and r['artifact_lifecycle_state']=='CURRENT' and r['controlling_published_baseline']==VERSION and r['baseline_date']==DATE for r in rows))
for rel in ['schemas/SCHEMA-CATALOG.md','schemas/README.md','api/INTERFACE-CATALOG.md','registries/REGISTRY-ENTRY-CATALOG.md','registries/README.md']:
 t=(ROOT/rel).read_text(); check('human_metadata_'+Path(rel).name,'CURRENT' in t and VERSION in t and STATUS in t and DATE in t,rel)
specs=['spec/AGCP-Multitenant-Operational-Specification.md','spec/ledger/AGCP-Append-Only-Governance-Ledger-Specification.md','spec/AGCP-Policy-Evaluation-Contract.md','spec/AGCP-Provenance-Wire-Format-Specification.md','spec/AGCP-Human-Review-Specification.md','spec/AGCP-Error-Mapping.md','spec/AGCP-HTTP-Interface-Specification.md']
for rel in specs:
 h='\n'.join((ROOT/rel).read_text().splitlines()[:20]); check('spec_header_'+Path(rel).stem,'Artifact Lifecycle:' in h and VERSION in h and 'Public Review Controlled Baseline' in h and DATE in h,h)
api=yaml.safe_load((ROOT/'api/AGCP-HTTP-Contract.yaml').read_text()); check('openapi_release_metadata',api['info']['version']==SPEC and api.get('x-agcp-specification-release')==VERSION)
man=yaml.safe_load((ROOT/'conformance/agcp-conformance-manifest.yml').read_text())['agcp_conformance_manifest']['spec']; check('manifest_release_metadata',man.get('repository_release_target')==VERSION and man.get('repository_release_target_status')=='public-review-controlled-baseline' and man.get('artifact_lifecycle_state')=='current' and man.get('controlling_published_baseline')==VERSION and man.get('controlling_baseline_status')=='public-review-controlled-baseline' and str(man.get('baseline_date'))==DATE and man.get('agcp_release')==VERSION,man)
wb=load_workbook(ROOT/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False); ws=wb[wb.sheetnames[0]]; hdr={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}; check('rtm_spec_versions',all(ws.cell(r,hdr['Specification_Version']).value=='v.2.0.4' for r in range(2,ws.max_row+1)))
for rel in ['registries/constraint-type-registry.json','registries/invariant-type-registry.json','registries/rejection-code-registry.json']:
 d=loadj(rel); x=copy.deepcopy(d); declared=x['integrity'].pop('document_digest')['value']; check('registry_digest_'+Path(rel).stem,declared==dig(x),{'declared':declared,'computed':dig(x)})
notes=(ROOT/'RELEASE_NOTES_v2.0.4.md').read_text(); check('release_notes_metadata',all(x in notes[:800] for x in ['Public Review Controlled Baseline','CURRENT','2026-08-05']))
report_files=['governance/AGCP-implementation-profile-validation.json','governance/AGCP-provenance-wire-format-validation.json','governance/AGCP-command-record-separation-validation.json','governance/AGCP-content-digest-contract-validation.json','governance/AGCP-http-error-metadata-validation.json','governance/AGCP-semantic-fixture-validation.json','governance/AGCP-normative-companion-reference-validation.json']
for rel in report_files:
 d=loadj(rel); check('report_release_context_'+Path(rel).stem,d.get('release_context')==ctx,d.get('release_context'))
source_files=[POLICY,'governance/AGCP-Release-Lifecycle-Metadata-Policy.md','schemas/catalog/schema-catalog.json','schemas/catalog/schema-catalog.csv','schemas/SCHEMA-CATALOG.md','schemas/README.md','api/interface-catalog.json','api/interface-catalog.csv','api/INTERFACE-CATALOG.md','registries/registry-entry-catalog.json','registries/registry-entry-catalog.csv','registries/REGISTRY-ENTRY-CATALOG.md','registries/constraint-type-registry.json','registries/invariant-type-registry.json','registries/rejection-code-registry.json','api/AGCP-HTTP-Contract.yaml','conformance/agcp-conformance-manifest.yml','RELEASE_NOTES_v2.0.4.md','governance/validate_release_lifecycle_metadata.py']+specs
report={'release_context':ctx,'report_id':'AGCP-P2-01-RELEASE-LIFECYCLE-METADATA','status':'PASS' if not issues else 'FAIL','validated_at':'2026-08-05','finding':'P2-01','catalog_versions':{'schema':'1.0.50','interface':'1.0.5','registry':'1.0.3'},'specification_count':len(specs),'validation_report_count':len(report_files),'source_hashes':{r:sha(r) for r in source_files},'checks':checks,'issues':issues}
(ROOT/'governance/AGCP-release-lifecycle-metadata-validation.json').write_text(json.dumps(report,indent=2,default=str)+'\n'); print(json.dumps({'status':report['status'],'checks':len(checks),'issues':issues},indent=2,default=str)); sys.exit(0 if not issues else 1)

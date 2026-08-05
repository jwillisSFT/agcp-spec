#!/usr/bin/env python3
"""Validate P1-12 algorithm-specific content-digest semantics across the AGCP repository."""
from __future__ import annotations
from pathlib import Path
import copy, hashlib, json, re, sys
import yaml
from jsonschema import Draft202012Validator, RefResolver
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
issues=[]; checks=[]
def loadj(r): return json.loads((ROOT/r).read_text())
def sha(r): return hashlib.sha256((ROOT/r).read_bytes()).hexdigest()
def check(name,ok,detail=''):
    checks.append({'check':name,'status':'PASS' if ok else 'FAIL','detail':detail})
    if not ok: issues.append(f'{name}: {detail}')
common=loadj('schemas/common.json'); cd=common['$defs']['content_digest']; hx=common['$defs']['hash_hex']
cv=Draft202012Validator(cd,resolver=RefResolver.from_schema(common)); hv=Draft202012Validator(hx,resolver=RefResolver.from_schema(common))
expected={'SHA-256':64,'SHA-384':96,'SHA-512':128,'BLAKE2B-256':64,'BLAKE2B-512':128}
branches={b['properties']['algorithm']['const']:b['properties']['value']['pattern'] for b in cd['oneOf']}
check('algorithm_branch_set',set(branches)==set(expected),branches)
check('algorithm_exact_patterns',all(branches[a]==f'^[0-9a-f]{{{n}}}$' for a,n in expected.items()),branches)
check('digest_closed_object',cd.get('additionalProperties') is False)
check('digest_required_fields',set(cd.get('required',[]))=={'algorithm','value'})
check('lowercase_base_pattern',cd['properties']['value'].get('pattern')=='^[0-9a-f]+$')
check('legacy_hash_hex_deprecated','Deprecated' in hx.get('description','') and len(hx.get('oneOf',[]))==3)
vec=loadj('conformance/digests/AGCP-Content-Digest-Test-Vectors.json')
pos=[{'id':v['id'],'valid':not list(cv.iter_errors(v['instance']))} for v in vec['positive_vectors']]
neg=[{'id':v['id'],'rejected':bool(list(cv.iter_errors(v['instance'])))} for v in vec['negative_vectors']]
check('positive_vector_count',len(pos)==5,len(pos));check('all_positive_vectors_valid',all(x['valid'] for x in pos),pos)
check('negative_vector_count',len(neg)==14,len(neg));check('all_negative_vectors_rejected',all(x['rejected'] for x in neg),neg)
ex=loadj('schemas/examples/common-content-digest-examples.json'); exr=[{'id':x['id'],'valid':not list(cv.iter_errors(x['digest']))} for x in ex['examples']]
check('controlled_example_count',len(exr)==5,len(exr));check('controlled_examples_valid',all(x['valid'] for x in exr),exr)
# Every direct schema dependent continues to reference the shared definition.
refs=[]
for p in sorted((ROOT/'schemas').glob('*.json')):
    if 'common.json#/$defs/content_digest' in p.read_text(): refs.append(str(p.relative_to(ROOT)))
check('dependent_schema_count',len(refs)==40,refs)
# Every algorithm/value digest object in controlled JSON/YAML data obeys the contract.
instances=[]; bad=[]
def walk(x,file,path='$'):
    if isinstance(x,dict):
        if x.get('algorithm') in expected and isinstance(x.get('value'),str):
            errs=[e.message for e in cv.iter_errors({k:v for k,v in x.items() if k in {'algorithm','value','canonicalization'}})]
            instances.append((file,path));
            if errs: bad.append({'file':file,'path':path,'errors':errs})
        for k,v in x.items(): walk(v,file,path+'.'+str(k))
    elif isinstance(x,list):
        for i,v in enumerate(x): walk(v,file,f'{path}[{i}]')
for base in ['schemas/examples','conformance']:
    for p in sorted((ROOT/base).rglob('*')):
        if not p.is_file() or p.suffix.lower() not in {'.json','.yaml','.yml'}: continue
        if p == ROOT/'conformance/digests/AGCP-Content-Digest-Test-Vectors.json': continue
        try: obj=json.loads(p.read_text()) if p.suffix.lower()=='.json' else yaml.safe_load(p.read_text())
        except Exception: continue
        walk(obj,str(p.relative_to(ROOT)))
check('controlled_digest_instances_valid',not bad,{'count':len(instances),'errors':bad})
# All mapped fixtures validate under their root schemas after the shared-definition change.
store={}
for p in (ROOT/'schemas').glob('*.json'):
    try:
        s=json.loads(p.read_text());store[s.get('$id',p.resolve().as_uri())]=s;Draft202012Validator.check_schema(s)
    except Exception as e: issues.append(f'{p.relative_to(ROOT)} metaschema: {e}')
fm=loadj('conformance/fixture-mapping.json'); fres=[]
for f in fm['fixtures']:
    s=loadj(f['schema_file']);o=loadj(f['example_file']);r=RefResolver(base_uri=(ROOT/f['schema_file']).resolve().as_uri(),referrer=s,store=store)
    errs=[e.message for e in Draft202012Validator(s,resolver=r).iter_errors(o)];actual='VALID' if not errs else 'INVALID'
    fres.append({'fixture_id':f['fixture_id'],'expected':f.get('expected_validation','VALID'),'actual':actual,'errors':errs[:3]})
check('controlled_fixtures_valid',len(fres)==30 and all(x['expected']==x['actual'] for x in fres),fres)
# OpenAPI and normative interface binding.
api=yaml.safe_load((ROOT/'api/AGCP-HTTP-Contract.yaml').read_text());comp=api['components']['schemas']['ContentDigest']
check('openapi_content_digest_ref',comp['allOf'][0]['$ref']=='../schemas/common.json#/$defs/content_digest')
check('openapi_content_digest_example',not list(cv.iter_errors(comp['example'])))
check('openapi_algorithm_length_map',comp['x-agcp-algorithm-output-lengths']==expected,comp['x-agcp-algorithm-output-lengths'])
ifs=(ROOT/'spec/AGCP-HTTP-Interface-Specification.md').read_text();check('if001_digest_rule','ambiguous identifier `BLAKE2B` SHALL fail' in ifs and '96 characters for `SHA-384`' in ifs)
# Catalog, RTM, and test mappings.
cat=loadj('schemas/catalog/schema-catalog.json');ds1=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-001')
check('schema_catalog_version',cat['catalog_version']=='1.0.50')
check('ds001_catalog_hash',ds1['sha256']==sha('schemas/common.json'),{'catalog':ds1['sha256'],'actual':sha('schemas/common.json')})
check('ds001_cr_mappings',set(['CR-042','CR-052','CR-064','CR-066']).issubset(ds1['cr_ids']),ds1['cr_ids'])
wb=load_workbook(ROOT/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False);ws=wb['AGCP_RTM_Repository_ARM_Co'];h={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)};rr=[]
for r in range(2,ws.max_row+1):
    cr=ws.cell(r,h['CR_ID']).value
    if cr in {'CR-042','CR-052','CR-064','CR-066'}:
        ds=str(ws.cell(r,h['DS_ID']).value or '').split(';');gf=str(ws.cell(r,h['GitHub_File']).value or '');nt=str(ws.cell(r,h['Notes']).value or '')
        rr.append({'cr':cr,'ds1':'DS-001' in ds,'files':all(x in gf for x in ['schemas/common.json','conformance/digests/AGCP-Content-Digest-Test-Vectors.json','governance/AGCP-content-digest-contract-validation.json']),'note':'P1-12 v2.0.1 correction' in nt})
check('rtm_digest_rows',len(rr)==4 and all(x['ds1'] and x['files'] and x['note'] for x in rr),rr)
tm=loadj('conformance/test-mapping.json');tr=[]
for t in tm['tests']:
    if t['tc_id'] in {'TC-042','TC-052','TC-064','TC-066'}:
        tr.append({'tc':t['tc_id'],'ds1':'DS-001' in t['ds_ids'],'common':'schemas/common.json' in t['schema_files'],'vectors':'conformance/digests/AGCP-Content-Digest-Test-Vectors.json' in t.get('supporting_companion_vectors',[]),'note':'p1_12_traceability_note' in t})
check('test_mapping_digest_rows',len(tr)==4 and all(all(v for k,v in x.items() if k!='tc') for x in tr),tr)
for rel,tc in [('conformance/tests/TC041-TC050.md','TC-042'),('conformance/tests/TC051-TC060.md','TC-052'),('conformance/tests/TC061-TC070.md','TC-064'),('conformance/tests/TC061-TC070.md','TC-066')]:
    check('formal_'+tc.lower().replace('-','_')+'_digest_clause','P1-12 Digest Contract:' in (ROOT/rel).read_text())
manifest=(ROOT/'conformance/agcp-conformance-manifest.yml').read_text();check('manifest_content_digest_assets',all(x in manifest for x in ['content_digest_contract:','conformance/digests/AGCP-Content-Digest-Test-Vectors.json','governance/validate_content_digest_contract.py']))
report={'release_context':{'repository_release_target':'v2.0.1','repository_release_target_status':'UNRELEASED_ACCUMULATED_CORRECTION_SET','controlling_published_baseline':'v2.0.0','controlling_baseline_status':'PUBLIC_REVIEW_CONTROLLED_BASELINE','baseline_date':'2026-07-30','artifact_lifecycle_state':'CURRENT'},'report_id':'AGCP-P1-12-CONTENT-DIGEST-CONTRACT','status':'PASS' if not issues else 'FAIL','validated_at':'2026-08-03','finding':'P1-12','schema':'schemas/common.json#/$defs/content_digest','algorithm_output_lengths':expected,'dependent_schema_count':len(refs),'controlled_digest_instance_count':len(instances),'positive_vector_count':len(pos),'negative_vector_count':len(neg),'affected_crs':['CR-042','CR-052','CR-064','CR-066'],'source_hashes':{r:sha(r) for r in ['schemas/common.json','schemas/examples/common-content-digest-examples.json','api/AGCP-HTTP-Contract.yaml','spec/AGCP-HTTP-Interface-Specification.md','conformance/digests/AGCP-Content-Digest-Test-Vectors.json','schemas/catalog/schema-catalog.json','conformance/test-mapping.json','spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx','governance/validate_content_digest_contract.py',]},'checks':checks,'issues':issues}
(ROOT/'governance/AGCP-content-digest-contract-validation.json').write_text(json.dumps(report,indent=2)+'\n')
print(json.dumps({'status':report['status'],'checks':len(checks),'issues':issues},indent=2));sys.exit(0 if not issues else 1)

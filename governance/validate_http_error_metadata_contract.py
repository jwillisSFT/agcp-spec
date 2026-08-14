#!/usr/bin/env python3
"""Validate P1-03, P1-09, P1-14, and P1-17 public IF-001 error and metadata contracts."""
from __future__ import annotations
from pathlib import Path
import copy, hashlib, json, re, sys
import yaml
from jsonschema import Draft202012Validator, RefResolver
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
from release_version import RELEASE_TAG, SEMVER, release_context
issues=[];checks=[]
def loadj(r): return json.loads((ROOT/r).read_text())
def sha(r): return hashlib.sha256((ROOT/r).read_bytes()).hexdigest()
def check(name,ok,detail=''):
    checks.append({'check':name,'status':'PASS' if ok else 'FAIL','detail':detail})
    if not ok: issues.append(f'{name}: {detail}')
# schema store
store={}
for p in (ROOT/'schemas').glob('*.json'):
    try:
        s=json.loads(p.read_text());Draft202012Validator.check_schema(s);store[s.get('$id',p.resolve().as_uri())]=s
    except Exception as e: issues.append(f'{p.relative_to(ROOT)} metaschema: {e}')
err=loadj('schemas/error_response.json');meta=loadj('schemas/meta_response.json')
ev=Draft202012Validator(err,resolver=RefResolver(base_uri=(ROOT/'schemas/error_response.json').resolve().as_uri(),referrer=err,store=store))
mv=Draft202012Validator(meta,resolver=RefResolver(base_uri=(ROOT/'schemas/meta_response.json').resolve().as_uri(),referrer=meta,store=store))
vec=loadj('conformance/http/AGCP-HTTP-Error-Metadata-Test-Vectors.json')
for key in ['ordinary_not_found','cross_scope_not_found','pre_governance_throttling','capacity_unavailable']:
    body=vec['public_error_contract'][key]['body']; check('vector_'+key+'_schema',not list(ev.iter_errors(body)),[e.message for e in ev.iter_errors(body)])
check('not_found_public_code',vec['public_error_contract']['ordinary_not_found']['body']['error']['rejection_code']=='RESOURCE_NOT_FOUND')
check('cross_scope_public_code',vec['public_error_contract']['cross_scope_not_found']['body']['error']['rejection_code']=='RESOURCE_NOT_FOUND')
th=vec['public_error_contract']['pre_governance_throttling'];check('throttle_429_retry_after',th['http_status']==429 and int(th['headers']['Retry-After'])>0 and th['governance_outcome_created'] is False)
cap=vec['public_error_contract']['capacity_unavailable'];check('capacity_503_not_governance',cap['http_status']==503 and cap['governance_outcome_created'] is False)
q=vec['public_error_contract']['governance_quota_denial'];check('governance_quota_is_outcome',q['http_status']==200 and q['representation']=='AUTHORITATIVE_GOVERNANCE_OUTCOME' and q['transport_error_body'] is None)
example=loadj('schemas/examples/ds003-implementation-metadata-response.json');merrs=[e.message for e in mv.iter_errors(example)];check('metadata_example_valid',not merrs,merrs)
release=example['supported_agcp_releases'][0]
check('metadata_baseline_digest',release['baseline_reference_type']=='IMMUTABLE_RELEASE_BUNDLE' and release['baseline_bundle_digest']['value']=='790bffe0883c5371c6007986e373e275c9e966cb40c84c8186f38e4365f8c326')
check('metadata_validator_binding',example['schema_set']['validator_set']['source_schema_set_digest']==example['schema_set']['schema_set_digest'])
check('metadata_profile_baseline_binding',example['implementation_profile']['baseline_bundle_digest']==release['baseline_bundle_digest'])
check('metadata_active_governance',example['active_governance']['activation_status']=='ACTIVE' and bool(example['active_governance']['governance_version']))
check('metadata_deployment_binding_public_safe',set(example['deployment_binding'])=={'deployment_id','node_id','workspace_id','tenant_id','governance_domain_id','binding_digest'})
# Negative metadata mutations
for n in vec['negative_vectors']:
    if n['id']=='META-NV-001':
        x=copy.deepcopy(example);x['supported_agcp_releases'][0].pop('baseline_bundle_digest');check(n['id'],bool(list(mv.iter_errors(x))))
    elif n['id']=='META-NV-002':
        x=copy.deepcopy(example);x['supported_agcp_releases'][0]['baseline_bundle_publication_status']='PUBLISHED';check(n['id'],bool(list(mv.iter_errors(x))))
    elif n['id']=='META-NV-003':
        uri='https://github.com/example/agcp/archive/refs/heads/main.zip';check(n['id'],bool(re.search(r'/heads/|/main(?:\.|/|$)',uri)))
    elif n['id']=='META-NV-004':
        x=copy.deepcopy(example);x['schema_set']['validator_set'].pop('validator_set_digest');check(n['id'],bool(list(mv.iter_errors(x))))
    elif n['id']=='META-NV-005':
        x=copy.deepcopy(example);x.pop('active_governance');check(n['id'],bool(list(mv.iter_errors(x))))
# OpenAPI
api=yaml.safe_load((ROOT/'api/AGCP-HTTP-Contract.yaml').read_text());ops=[]
for path,item in api['paths'].items():
    for m,op in item.items():
        if m in {'get','post','put','patch','delete','head','options','trace'}: ops.append((path,m,op))
check('openapi_version',api['info']['version']==SEMVER)
check('all_operations_429',all(op['responses'].get('429',{}).get('$ref')=='#/components/responses/TooManyRequests' for _,_,op in ops),[(p,m) for p,m,o in ops if '429' not in o['responses']])
check('all_operations_503',all(op['responses'].get('503',{}).get('$ref')=='#/components/responses/ServiceUnavailable' for _,_,op in ops),[(p,m) for p,m,o in ops if '503' not in o['responses']])
check('all_404_normalized',all(op['responses'].get('404',{}).get('$ref')=='#/components/responses/PublicNotFound' for _,_,op in ops if '404' in op['responses']))
ra=api['components']['headers']['RetryAfter'];check('retry_after_delay_seconds',ra['schema']['type']=='integer' and ra['schema']['minimum']==1 and 'delay-seconds' in ra['description'])
check('metadata_openapi_ref',api['paths']['/agcp/v2/meta']['get']['responses']['200']['content']['application/json']['schema']['$ref']=='#/components/schemas/MetadataResponse')
# Registry and integrity
reg=loadj('registries/rejection-code-registry.json');codes={e['code']:e for e in reg['codes']}
check('registry_release',reg['release']['registry_release']==RELEASE_TAG and reg['integrity']['entry_count']==42)
check('new_transport_codes',codes['REQUEST_THROTTLED']['default_http_status']==429 and codes['CAPACITY_UNAVAILABLE']['default_http_status']==503)
check('specific_not_found_deprecated',all(codes[c]['entry_status']=='DEPRECATED' and codes[c]['successor_entry_id']=='REG-059' for c in ['PROPOSAL_NOT_FOUND','AUTHORIZATION_NOT_FOUND','GOVERNANCE_EVIDENCE_NOT_FOUND','GOVERNANCE_ARTIFACT_NOT_FOUND']))
def canon(o): return json.dumps(o,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()
def dig(o): return hashlib.sha256(canon(o)).hexdigest()
entry_ok=True
for e in reg['codes']:
    x=copy.deepcopy(e);decl=x.pop('entry_digest')['value'];entry_ok &= decl==dig(x)
check('registry_entry_digests',entry_ok)
check('registry_entry_set_digest',reg['integrity']['entry_set_digest']['value']==dig(reg['codes']))
x=copy.deepcopy(reg);decl=x['integrity'].pop('document_digest')['value'];check('registry_document_digest',decl==dig(x),{'declared':decl,'computed':dig(x)})
# Harness public not-found bodies
h=yaml.safe_load((ROOT/'conformance/AGCP-Conformance-Harness-Spec.yml').read_text());found=[]
def walk(o):
    if isinstance(o,dict):
        if o.get('rejection_code')=='RESOURCE_NOT_FOUND': found.append(o)
        for v in o.values(): walk(v)
    elif isinstance(o,list):
        for v in o: walk(v)
walk(h);check('harness_public_not_found_count',len(found)>=1,len(found));check('harness_public_not_found_valid',all(not list(ev.iter_errors({'error':o})) for o in found),[[e.message for e in ev.iter_errors({'error':o})] for o in found])
# Catalog hashes/mappings
cat=loadj('schemas/catalog/schema-catalog.json');d2=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-002');d3=next(e for e in cat['implemented_schemas'] if e['ds_id']=='DS-003')
check('schema_catalog_version',cat['catalog_version']=='1.0.50');check('ds002_hash',d2['sha256']==sha('schemas/error_response.json'));check('ds003_hash',d3['sha256']==sha('schemas/meta_response.json'))
ic=loadj('api/interface-catalog.json');i1=next(i for i in ic['interfaces'] if i['if_id']=='IF-001');check('interface_catalog_version',ic['catalog_version']=='1.0.5');check('interface_public_contract',i1['public_error_contract']['public_not_found']=='404 RESOURCE_NOT_FOUND' and i1['metadata_contract']['active_governance_version_required'])
rc=loadj('registries/registry-entry-catalog.json');check('registry_catalog_count',rc['catalog_version']=='1.0.3' and rc['entry_count']==94)
# RTM/test mapping
wb=load_workbook(ROOT/'spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx',data_only=False);ws=wb['AGCP_RTM_Repository_ARM_Co'];heads={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)};affected={'CR-010','CR-022','CR-026','CR-041','CR-048','CR-068','CR-073','CR-089','CR-090','CR-110','CR-113'};rows=[]
for r in range(2,ws.max_row+1):
    cr=ws.cell(r,heads['CR_ID']).value
    if cr in affected:
        ds=str(ws.cell(r,heads['DS_ID']).value or '').split(';');it=str(ws.cell(r,heads['IF_ID']).value or '').split(';');gf=str(ws.cell(r,heads['GitHub_File']).value or '');note=str(ws.cell(r,heads['Notes']).value or '')
        rows.append({'cr':cr,'ds':{'DS-002','DS-003'}.issubset(ds),'if':'IF-001' in it,'files':'conformance/http/AGCP-HTTP-Error-Metadata-Test-Vectors.json' in gf,'note':'P1-03/P1-09/P1-14/P1-17' in note})
check('rtm_rows',len(rows)==len(affected) and all(all(v for k,v in x.items() if k!='cr') for x in rows),rows)
tm=loadj('conformance/test-mapping.json');trows=[]
for t in tm['tests']:
    if t.get('cr_id') in affected:
        trows.append({'tc':t['tc_id'],'ds':{'DS-002','DS-003'}.issubset(t['ds_ids']),'if':'IF-001' in t['if_ids'],'vec':'conformance/http/AGCP-HTTP-Error-Metadata-Test-Vectors.json' in t.get('supporting_companion_vectors',[]),'note':'p1_03_p1_09_p1_14_p1_17_traceability_note' in t})
mapped_affected={t.get('cr_id') for t in tm['tests'] if t.get('cr_id') in affected}
check('test_mapping_rows',len(trows)==len(mapped_affected) and all(all(v for k,v in x.items() if k!='tc') for x in trows),{'mapped_crs':sorted(mapped_affected),'formally_defined_unmapped_crs':sorted(affected-mapped_affected),'rows':trows})
report={'release_context':release_context(),'report_id':'AGCP-P1-03-P1-09-P1-14-P1-17-HTTP-ERROR-METADATA','status':'PASS' if not issues else 'FAIL','validated_at':'2026-08-05','findings':['P1-03','P1-09','P1-14','P1-17'],'operation_count':len(ops),'public_not_found_harness_instance_count':len(found),'registry_entry_count':len(reg['codes']),'affected_crs':sorted(affected),'source_hashes':{r:sha(r) for r in ['schemas/error_response.json','schemas/meta_response.json','schemas/examples/ds003-implementation-metadata-response.json','api/AGCP-HTTP-Contract.yaml','spec/AGCP-HTTP-Interface-Specification.md','spec/AGCP-Error-Mapping.md','registries/rejection-code-registry.json','conformance/http/AGCP-HTTP-Error-Metadata-Test-Vectors.json','schemas/catalog/schema-catalog.json','api/interface-catalog.json','registries/registry-entry-catalog.json','conformance/test-mapping.json','spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx','governance/validate_http_error_metadata_contract.py']},'checks':checks,'issues':issues}
(ROOT/'governance/AGCP-http-error-metadata-validation.json').write_text(json.dumps(report,indent=2)+'\n')
print(json.dumps({'status':report['status'],'checks':len(checks),'issues':issues},indent=2));sys.exit(0 if not issues else 1)

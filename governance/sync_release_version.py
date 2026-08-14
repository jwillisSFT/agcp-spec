#!/usr/bin/env python3
"""Synchronize current AGCP release metadata from the root VERSION file.

The root VERSION file is the sole maintained repository release number.
This synchronizer updates current-release metadata only. Historical release
notes and release-specific historical validation/synchronization records are
not rewritten.
"""
from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import json
import re
from pathlib import Path

import yaml
from openpyxl import load_workbook

from release_version import (
    ROOT,
    SEMVER,
    RELEASE_TAG,
    RTM_SPEC_VERSION,
    RELEASE_IDENTIFIER,
    CURRENT_RELEASE_NOTES,
    RELEASE_STATUS,
    BASELINE_DATE,
    ARTIFACT_LIFECYCLE_STATE,
    SETTINGS,
    SYNC_MANIFEST,
    SYNC_REPORT,
    INTEGRITY_REPORT,
    release_context,
    release_lifecycle_metadata,
)

REPORT = ROOT / "governance/AGCP-version-source-validation.json"

CURRENT_CATALOG_JSON = [
    "schemas/catalog/schema-catalog.json",
    "api/interface-catalog.json",
    "registries/registry-entry-catalog.json",
]

CURRENT_CATALOG_CSV = [
    "schemas/catalog/schema-catalog.csv",
    "api/interface-catalog.csv",
    "registries/registry-entry-catalog.csv",
]

CURRENT_REGISTRIES = [
    "registries/constraint-type-registry.json",
    "registries/invariant-type-registry.json",
    "registries/rejection-code-registry.json",
]

CURRENT_RELEASE_METADATA_JSON = [
    "governance/normative-companion-reference-dispositions.json",
]

CURRENT_HUMAN_METADATA = [
    "schemas/SCHEMA-CATALOG.md",
    "schemas/README.md",
    "api/INTERFACE-CATALOG.md",
    "registries/REGISTRY-ENTRY-CATALOG.md",
    "registries/README.md",
    "governance/AGCP-Release-Lifecycle-Metadata-Policy.md",
    "governance/AGCP-Normative-Companion-Reference-Dispositions.md",
    "spec/AGCP-Multitenant-Operational-Specification.md",
    "spec/ledger/AGCP-Append-Only-Governance-Ledger-Specification.md",
    "spec/AGCP-Policy-Evaluation-Contract.md",
    "spec/AGCP-Provenance-Wire-Format-Specification.md",
    "spec/AGCP-Human-Review-Specification.md",
    "spec/AGCP-Error-Mapping.md",
    "spec/AGCP-HTTP-Interface-Specification.md",
    "implementer/IMPLEMENTATION-PROFILE-CATALOG.md",
    "implementer/README.md",
    "spec/README.md",
    "conformance/README.md",
    "conformance/AGCP-Conformance-Traceability-and-Automation-Model.md",
    "ARCHITECTURE.md",
]


def write_text_if_changed(path: Path, text: str, write: bool, changed: list[str]) -> None:
    old = path.read_text(encoding="utf-8") if path.exists() else None
    if old == text:
        return
    changed.append(path.relative_to(ROOT).as_posix())
    if write:
        path.write_text(text, encoding="utf-8")


def write_json_if_changed(path: Path, value: object, write: bool, changed: list[str]) -> None:
    desired = json.dumps(value, indent=2, ensure_ascii=False) + "\n"
    write_text_if_changed(path, desired, write, changed)


def canonical_digest(value: object) -> str:
    payload = json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode()
    return hashlib.sha256(payload).hexdigest()


def sync_machine_release_metadata(write: bool, changed: list[str]) -> None:
    # Machine-readable lifecycle metadata is generated from VERSION.
    meta_path = ROOT / "governance/release-lifecycle-metadata.json"
    write_json_if_changed(meta_path, release_lifecycle_metadata(), write, changed)

    # OpenAPI current release metadata.
    api_path = ROOT / "api/AGCP-HTTP-Contract.yaml"
    api_text = api_path.read_text(encoding="utf-8")
    updated = re.sub(r"(?m)^(  version:)\s*[^\n]+$", rf"\1 {SEMVER}", api_text, count=1)
    updated = re.sub(r"(?m)^(x-agcp-specification-release:)\s*[^\n]+$", rf"\1 {RELEASE_TAG}", updated, count=1)
    updated = re.sub(r"AGCP v\d+\.\d+\.\d+ contract-compatible", f"AGCP {RELEASE_TAG} contract-compatible", updated)
    write_text_if_changed(api_path, updated, write, changed)

    # Conformance manifest release identity.
    man_path = ROOT / "conformance/agcp-conformance-manifest.yml"
    man_text = man_path.read_text(encoding="utf-8")
    replacements = {
        "repository_release_target": RELEASE_TAG,
        "repository_release_target_status": RELEASE_STATUS.lower().replace("_", "-"),
        "artifact_lifecycle_state": ARTIFACT_LIFECYCLE_STATE.lower(),
        "controlling_published_baseline": RELEASE_TAG,
        "controlling_baseline_status": RELEASE_STATUS.lower().replace("_", "-"),
        "baseline_date": BASELINE_DATE,
        "agcp_release": RELEASE_TAG,
    }
    lines = man_text.splitlines(True)
    in_spec = False
    for i, line in enumerate(lines):
        if line.startswith("  spec:"):
            in_spec = True
            continue
        if in_spec and line.startswith("  ") and not line.startswith("    "):
            break
        if in_spec:
            match = re.match(r"(    ([A-Za-z0-9_]+):)\s*.*?(\r?\n)?$", line)
            if match and match.group(2) in replacements:
                nl = match.group(3) or ""
                lines[i] = f"    {match.group(2)}: {replacements[match.group(2)]}{nl}"
    updated = "".join(lines)
    if not re.search(r"(?m)^    version_source: VERSION$", updated):
        updated = updated.replace(
            "    release_model: repository_release\n",
            "    release_model: repository_release\n    version_source: VERSION\n",
            1,
        )
    write_text_if_changed(man_path, updated, write, changed)

    # Current catalog JSON metadata.
    for rel in CURRENT_CATALOG_JSON:
        path = ROOT / rel
        obj = json.loads(path.read_text(encoding="utf-8"))
        for key, value in {
            "specification_version": RELEASE_TAG,
            "repository_release_target": RELEASE_TAG,
            "repository_release_target_status": RELEASE_STATUS,
            "release_status": RELEASE_STATUS,
            "controlling_published_baseline": RELEASE_TAG,
            "baseline_date": BASELINE_DATE,
            "artifact_lifecycle_state": ARTIFACT_LIFECYCLE_STATE,
        }.items():
            if key in obj:
                obj[key] = value
        if rel == "schemas/catalog/schema-catalog.json":
            for entry in obj.get("implemented_schemas", []) + obj.get("retired_schemas", []):
                if "specification_version" in entry:
                    entry["specification_version"] = RELEASE_TAG
        if rel == "api/interface-catalog.json":
            for entry in obj.get("interfaces", []):
                if "contract_version" in entry:
                    entry["contract_version"] = SEMVER
        write_json_if_changed(path, obj, write, changed)

    # Implementation Profile catalog/manifest release identity.
    profile_catalog = ROOT / "implementer/implementation-profile-catalog.json"
    obj = json.loads(profile_catalog.read_text(encoding="utf-8"))
    obj["release_target"] = SEMVER
    for key, value in {
        "repository_release_target": RELEASE_TAG,
        "repository_release_target_status": RELEASE_STATUS,
        "controlling_published_baseline": RELEASE_TAG,
        "baseline_date": BASELINE_DATE,
        "artifact_lifecycle_state": ARTIFACT_LIFECYCLE_STATE,
    }.items():
        if key in obj:
            obj[key] = value
    write_json_if_changed(profile_catalog, obj, write, changed)

    # Fixture mapping identifies the current specification release.
    fixture_path = ROOT / "conformance/fixture-mapping.json"
    fixture = json.loads(fixture_path.read_text(encoding="utf-8"))
    if "specification_version" in fixture:
        fixture["specification_version"] = RELEASE_TAG
    write_json_if_changed(fixture_path, fixture, write, changed)

    # Test mapping points at current VERSION-derived synchronization artifacts.
    test_map_path = ROOT / "conformance/test-mapping.json"
    test_map = json.loads(test_map_path.read_text(encoding="utf-8"))
    sync = test_map.setdefault("repository_synchronization", {})
    sync["manifest"] = SYNC_MANIFEST
    sync["validation_report"] = SYNC_REPORT
    write_json_if_changed(test_map_path, test_map, write, changed)

    # Additional current machine-readable release metadata.
    for rel in CURRENT_RELEASE_METADATA_JSON:
        path = ROOT / rel
        obj = json.loads(path.read_text(encoding="utf-8"))
        if rel == "governance/normative-companion-reference-dispositions.json":
            obj["release_target"] = RELEASE_TAG
        write_json_if_changed(path, obj, write, changed)

    # Current conformance-harness metadata is governed by the repository release.
    # Version values inside semantic fixture/test payloads remain untouched because
    # those values model data, not repository release authority.
    harness_spec_path = ROOT / "conformance/AGCP-Conformance-Harness-Spec.yml"
    harness_spec_text = harness_spec_path.read_text(encoding="utf-8")
    harness_spec_updated = re.sub(
        r"(?m)^(  agcp_release:)\s*v\d+\.\d+\.\d+$",
        rf"\1 {RELEASE_TAG}",
        harness_spec_text,
        count=1,
    )
    write_text_if_changed(harness_spec_path, harness_spec_updated, write, changed)

    harness_checks_path = ROOT / "conformance/harness-checks.json"
    harness_checks = json.loads(harness_checks_path.read_text(encoding="utf-8"))
    harness_checks["agcp_release"] = RELEASE_TAG
    write_json_if_changed(harness_checks_path, harness_checks, write, changed)

    # Current controlled registries use current release identity. Historical entry
    # introduction metadata is left untouched. Recompute document digests after
    # changing the registry release envelope.
    for rel in CURRENT_REGISTRIES:
        path = ROOT / rel
        obj = json.loads(path.read_text(encoding="utf-8"))
        release = obj.setdefault("release", {})
        old_release_id = release.get("release_id")
        release["release_id"] = f"{obj.get('registry_id', 'REGISTRY')}-{RELEASE_TAG}"
        release["registry_release"] = RELEASE_TAG
        release["specification_version"] = RELEASE_TAG
        release["publication_status"] = "CURRENT"
        if old_release_id and old_release_id != release["release_id"] and "supersedes_release_id" not in release:
            release["supersedes_release_id"] = old_release_id
        digest_source = copy.deepcopy(obj)
        digest_source.get("integrity", {}).pop("document_digest", None)
        obj["integrity"]["document_digest"] = {
            "algorithm": "SHA-256",
            "value": canonical_digest(digest_source),
        }
        write_json_if_changed(path, obj, write, changed)


def sync_catalog_csv_metadata(write: bool, changed: list[str]) -> None:
    for rel in CURRENT_CATALOG_CSV:
        path = ROOT / rel
        with path.open(newline="", encoding="utf-8-sig") as handle:
            rows = list(csv.DictReader(handle))
        if not rows:
            continue
        fields = list(rows[0].keys())
        dirty = False
        for row in rows:
            for key, value in {
                "repository_release_target": RELEASE_TAG,
                "repository_release_target_status": RELEASE_STATUS,
                "release_status": RELEASE_STATUS,
                "artifact_lifecycle_state": ARTIFACT_LIFECYCLE_STATE,
                "controlling_published_baseline": RELEASE_TAG,
                "baseline_date": BASELINE_DATE,
            }.items():
                if key in row and row[key] != value:
                    row[key] = value
                    dirty = True
            if rel.startswith("schemas/") and "specification_version" in row and row["specification_version"] != RELEASE_TAG:
                row["specification_version"] = RELEASE_TAG
                dirty = True
            if rel.startswith("api/") and "contract_version" in row and row["contract_version"] != SEMVER:
                row["contract_version"] = SEMVER
                dirty = True
        if dirty:
            changed.append(rel)
            if write:
                with path.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(handle, fieldnames=fields)
                    writer.writeheader()
                    writer.writerows(rows)


def replace_current_metadata_labels(text: str) -> str:
    # Only controlled current-release metadata labels are rewritten. Historical
    # prose such as "resolved in v2.0.4" is intentionally left unchanged.
    patterns = [
        (r"(?mi)(Specification Version:\*\*\s*`?)(?:v)?\d+\.\d+\.\d+(`?)", rf"\g<1>{SEMVER}\2"),
        (r"(?mi)(Specification version:\s*`?)v\d+\.\d+\.\d+(`?)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(AGCP specification version:\s*`?)v\d+\.\d+\.\d+(`?)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(AGCP Specification Release:\*\*\s*)v\d+\.\d+\.\d+", rf"\g<1>{RELEASE_TAG}"),
        (r"(?mi)(Contract Version:\*\*\s*)\d+\.\d+\.\d+", rf"\g<1>{SEMVER}"),
        (r"(?mi)(Repository Release Target:\*\*\s*AGCP\s+)v\d+\.\d+\.\d+", rf"\g<1>{RELEASE_TAG}"),
        (r"(?mi)(Repository release target\s*\|\s*`)v\d+\.\d+\.\d+(`)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(Repository release target:\s*`?)v\d+\.\d+\.\d+(`?)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(Controlling Published Baseline:\*\*\s*AGCP\s+)v\d+\.\d+\.\d+", rf"\g<1>{RELEASE_TAG}"),
        (r"(?mi)(Controlling published baseline\s*\|\s*`)v\d+\.\d+\.\d+(`)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(Controlling published baseline:\s*`?)v\d+\.\d+\.\d+(`?)", rf"\g<1>{RELEASE_TAG}\2"),
        (r"(?mi)(Release target:\*\*\s*AGCP\s+)v\d+\.\d+\.\d+", rf"\g<1>{RELEASE_TAG}"),
        (r"(?mi)(Baseline Date:\*\*\s*)\d{4}-\d{2}-\d{2}", rf"\g<1>{BASELINE_DATE}"),
        (r"(?mi)(Baseline date:\s*`?)\d{4}-\d{2}-\d{2}(`?)", rf"\g<1>{BASELINE_DATE}\2"),
        (r"(?mi)(Baseline date\s*\|\s*`?)\d{4}-\d{2}-\d{2}(`?)", rf"\g<1>{BASELINE_DATE}\2"),
    ]
    for pattern, replacement in patterns:
        text = re.sub(pattern, replacement, text)
    return text


def sync_human_metadata(write: bool, changed: list[str]) -> None:
    for rel in CURRENT_HUMAN_METADATA:
        path = ROOT / rel
        text = path.read_text(encoding="utf-8")
        updated = replace_current_metadata_labels(text)
        if rel == "api/INTERFACE-CATALOG.md":
            updated = re.sub(
                r"(\| `IF-00[12]` \|[^\n]*?\| `v2` \| `)\d+\.\d+\.\d+(` \|)",
                rf"\g<1>{SEMVER}\2",
                updated,
            )
        elif rel == "schemas/README.md":
            updated = re.sub(
                r"(authoritative JSON Schema Draft 2020-12 definitions for the AGCP )v\d+\.\d+\.\d+( schema set)",
                rf"\g<1>{RELEASE_TAG}\2",
                updated,
                count=1,
            )
        elif rel == "governance/AGCP-Release-Lifecycle-Metadata-Policy.md":
            updated = re.sub(
                r"(Current repository specification version \| `)\d+\.\d+\.\d+(`)",
                rf"\g<1>{SEMVER}\2",
                updated,
            )
            updated = re.sub(
                r"(identify the same `)v\d+\.\d+\.\d+(` artifact set)",
                rf"\g<1>{RELEASE_TAG}\2",
                updated,
            )
        elif rel == "spec/AGCP-HTTP-Interface-Specification.md":
            updated = re.sub(
                r"The current contract revision is v\d+\.\d+\.\d+ and is part of the published v\d+\.\d+\.\d+ Public Review Controlled Baseline\.",
                f"The current contract revision is {RELEASE_TAG} and is part of the published {RELEASE_TAG} Public Review Controlled Baseline.",
                updated,
            )
        elif rel == "ARCHITECTURE.md":
            updated = re.sub(
                r"`governance/AGCP-v\d+\.\d+\.\d+-repository-synchronization-manifest\.json`",
                f"`{SYNC_MANIFEST}`",
                updated,
            )
        elif rel == "conformance/README.md":
            updated = re.sub(
                r"The cumulative v\d+\.\d+\.\d+ (?:correction|repository) set is indexed by `\.\./governance/AGCP-v\d+\.\d+\.\d+-repository-synchronization-manifest\.json`",
                f"The cumulative {RELEASE_TAG} repository set is indexed by `../{SYNC_MANIFEST}`",
                updated,
            )
        elif rel == "conformance/AGCP-Conformance-Traceability-and-Automation-Model.md":
            updated = re.sub(
                r"(controlled relationships for the current AGCP )v\d+\.\d+\.\d+( model are)",
                rf"\g<1>{RELEASE_TAG}\2",
                updated,
            )
            try:
                harness_count = len(json.loads((ROOT / "conformance/harness-checks.json").read_text(encoding="utf-8")).get("checks", []))
                harness_spec = yaml.safe_load((ROOT / "conformance/AGCP-Conformance-Harness-Spec.yml").read_text(encoding="utf-8")) or {}
                vector_count = len(harness_spec.get("tests", []))
                updated = re.sub(
                    r"At the controlled AGCP v\d+\.\d+\.\d+ baseline, the mapping set contains 122 Formal Test Cases, \d+ Harness Checks, and \d+ Harness Test Vectors\.",
                    f"At the controlled AGCP {RELEASE_TAG} baseline, the mapping set contains 122 Formal Test Cases, {harness_count} Harness Checks, and {vector_count} Harness Test Vectors.",
                    updated,
                )
            except Exception:
                pass
        write_text_if_changed(path, updated, write, changed)

    # Root README contains current-release prose in addition to metadata labels.
    path = ROOT / "README.md"
    text = path.read_text(encoding="utf-8")
    updated = replace_current_metadata_labels(text)
    updated = re.sub(r"(?m)^> \*\*Current release:\*\* AGCP v\d+\.\d+\.\d+", f"> **Current release:** AGCP {RELEASE_TAG}", updated)
    updated = re.sub(r"(?m)^> \*\*Controlling published baseline:\*\* AGCP v\d+\.\d+\.\d+", f"> **Controlling published baseline:** AGCP {RELEASE_TAG}", updated)
    updated = re.sub(r"(?m)^> This repository snapshot is the controlled AGCP v\d+\.\d+\.\d+", f"> This repository snapshot is the controlled AGCP {RELEASE_TAG}", updated)
    updated = re.sub(r"(?m)^> \*\*Current release notes:\*\* .*?$", f"> **Current release notes:** [`{CURRENT_RELEASE_NOTES}`]({CURRENT_RELEASE_NOTES})  ", updated)
    updated = re.sub(
        r"For this repository snapshot, `VERSION` is `\d+\.\d+\.\d+`, yielding release tag `v\d+\.\d+\.\d+` and RTM specification version `v\.\d+\.\d+\.\d+`\.",
        f"For this repository snapshot, `VERSION` is `{SEMVER}`, yielding release tag `{RELEASE_TAG}` and RTM specification version `{RTM_SPEC_VERSION}`.",
        updated,
    )
    updated = re.sub(r"The v\d+\.\d+\.\d+ controlled inventory contains \*\*357 unique Normative Statement identifiers\*\*\.", f"The {RELEASE_TAG} controlled inventory contains **357 unique Normative Statement identifiers**.", updated, count=1)
    updated = re.sub(r"AGCP v\d+\.\d+\.\d+ is issued as the controlled \*\*Public Review Controlled Baseline\*\*", f"AGCP {RELEASE_TAG} is issued as the controlled **Public Review Controlled Baseline**", updated, count=1)
    updated = re.sub(r"This controlled release is AGCP v\d+\.\d+\.\d+ Public Review Controlled Baseline", f"This controlled release is AGCP {RELEASE_TAG} Public Review Controlled Baseline", updated, count=1)
    updated = re.sub(r"AGCP v\d+\.\d+\.\d+ is the current Public Review Controlled Baseline\.", f"AGCP {RELEASE_TAG} is the current Public Review Controlled Baseline.", updated, count=1)
    updated = re.sub(r"(Public Review Controlled Baseline\*\* dated )\d{4}-\d{2}-\d{2}", rf"\g<1>{BASELINE_DATE}", updated, count=1)
    updated = re.sub(r"(Public Review Controlled Baseline, baseline date )\d{4}-\d{2}-\d{2}", rf"\g<1>{BASELINE_DATE}", updated, count=1)
    updated = re.sub(r"(controlled baseline date is `)\d{4}-\d{2}-\d{2}(`)", rf"\g<1>{BASELINE_DATE}\2", updated, count=1)
    updated = re.sub(r"## v\d+\.\d+\.\d+ repository-wide integrity gate", f"## {RELEASE_TAG} repository-wide integrity gate", updated)
    updated = re.sub(r"`governance/AGCP-v\d+\.\d+\.\d+-repository-integrity-validation\.json`", f"`{INTEGRITY_REPORT}`", updated)
    updated = re.sub(r"## v\d+\.\d+\.\d+ repository synchronization", f"## {RELEASE_TAG} repository synchronization", updated)
    updated = re.sub(
        r"(?m)^├── RELEASE_NOTES_v\d+\.\d+\.\d+\.md$",
        f"├── {CURRENT_RELEASE_NOTES}",
        updated,
        count=1,
    )
    updated = re.sub(
        r"(?m)^- `RELEASE_NOTES_v\d+\.\d+\.\d+\.md` — release notes for the current v\d+\.\d+\.\d+ Public Review Controlled Baseline$",
        f"- `{CURRENT_RELEASE_NOTES}` — release notes for the current {RELEASE_TAG} Public Review Controlled Baseline",
        updated,
        count=1,
    )
    write_text_if_changed(path, updated, write, changed)


def sync_rtm(write: bool, changed: list[str]) -> None:
    path = ROOT / "spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx"
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    header = {worksheet.cell(1, col).value: col for col in range(1, worksheet.max_column + 1)}
    modified = False
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, header["Specification_Version"]).value != RTM_SPEC_VERSION:
            worksheet.cell(row, header["Specification_Version"]).value = RTM_SPEC_VERSION
            modified = True
    if "NS_Inventory_Dispositions" in workbook.sheetnames:
        sheet = workbook["NS_Inventory_Dispositions"]
        disposition_header = {sheet.cell(1, col).value: col for col in range(1, sheet.max_column + 1)}
        if "Specification_Version" in disposition_header:
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, disposition_header["Specification_Version"]).value != RTM_SPEC_VERSION:
                    sheet.cell(row, disposition_header["Specification_Version"]).value = RTM_SPEC_VERSION
                    modified = True
    if modified:
        changed.append(path.relative_to(ROOT).as_posix())
        if write:
            workbook.save(path)


def refresh_profile_manifest(write: bool, changed: list[str]) -> None:
    path = ROOT / "implementer/implementation-profile-manifest.json"
    obj = json.loads(path.read_text(encoding="utf-8"))
    obj["release_target"] = RELEASE_TAG
    for entry in obj.get("files", []):
        file_path = ROOT / "implementer" / entry["path"]
        if file_path.is_file():
            entry["sha256"] = hashlib.sha256(file_path.read_bytes()).hexdigest()
            entry["bytes"] = file_path.stat().st_size
    write_json_if_changed(path, obj, write, changed)


def sync(write: bool) -> tuple[list[str], list[str]]:
    changed: list[str] = []
    issues: list[str] = []

    sync_machine_release_metadata(write, changed)
    sync_catalog_csv_metadata(write, changed)
    sync_human_metadata(write, changed)
    sync_rtm(write, changed)
    # Profile package hashes depend on human and machine catalog content.
    refresh_profile_manifest(write, changed)

    notes = ROOT / CURRENT_RELEASE_NOTES
    if not notes.is_file():
        issues.append(f"missing current release notes: {CURRENT_RELEASE_NOTES}")
    elif RELEASE_TAG not in notes.read_text(encoding="utf-8")[:1200]:
        issues.append(f"current release notes header does not identify {RELEASE_TAG}")

    return changed, issues


def validate() -> tuple[list[dict], list[str]]:
    checks: list[dict] = []
    issues: list[str] = []

    def ck(name: str, ok: bool, detail=None) -> None:
        checks.append({"check": name, "status": "PASS" if ok else "FAIL", "detail": detail})
        if not ok:
            issues.append(f"{name}: {detail}")

    meta = json.loads((ROOT / "governance/release-lifecycle-metadata.json").read_text(encoding="utf-8"))
    ck(
        "authoritative_version_source",
        (ROOT / "VERSION").read_text().strip() == SEMVER and meta.get("version_source") == "VERSION",
        {"VERSION": SEMVER, "metadata_version_source": meta.get("version_source")},
    )
    ck(
        "release_metadata_generated_from_version",
        meta.get("repository_release_target") == RELEASE_TAG
        and meta.get("current_repository_specification_version") == SEMVER
        and meta.get("controlling_published_baseline", {}).get("release_id") == RELEASE_IDENTIFIER,
        meta,
    )

    api = yaml.safe_load((ROOT / "api/AGCP-HTTP-Contract.yaml").read_text(encoding="utf-8"))
    ck(
        "openapi_version",
        api.get("info", {}).get("version") == SEMVER and api.get("x-agcp-specification-release") == RELEASE_TAG,
        {"info.version": api.get("info", {}).get("version"), "release": api.get("x-agcp-specification-release")},
    )

    manifest = yaml.safe_load((ROOT / "conformance/agcp-conformance-manifest.yml").read_text(encoding="utf-8"))["agcp_conformance_manifest"]["spec"]
    ck(
        "conformance_manifest_version",
        manifest.get("version_source") == "VERSION"
        and manifest.get("repository_release_target") == RELEASE_TAG
        and manifest.get("controlling_published_baseline") == RELEASE_TAG
        and manifest.get("agcp_release") == RELEASE_TAG,
        manifest,
    )

    for rel in CURRENT_CATALOG_JSON:
        obj = json.loads((ROOT / rel).read_text(encoding="utf-8"))
        ck(
            "catalog_release_metadata_" + Path(rel).stem,
            obj.get("specification_version") == RELEASE_TAG
            and obj.get("repository_release_target") == RELEASE_TAG
            and obj.get("controlling_published_baseline") == RELEASE_TAG,
            {k: obj.get(k) for k in ["specification_version", "repository_release_target", "controlling_published_baseline"]},
        )

    workbook = load_workbook(ROOT / "spec/AGCP_Requirements_Traceability_Matrix_(RTM).xlsx", data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    header = {worksheet.cell(1, col).value: col for col in range(1, worksheet.max_column + 1)}
    values = {worksheet.cell(row, header["Specification_Version"]).value for row in range(2, worksheet.max_row + 1)}
    ck("rtm_specification_version", values == {RTM_SPEC_VERSION}, sorted(str(value) for value in values))

    ck("current_release_notes_present", (ROOT / CURRENT_RELEASE_NOTES).is_file(), CURRENT_RELEASE_NOTES)

    # Executable code may import VERSION-derived constants but must not embed the
    # current repository semantic version as an independent literal dependency.
    # Historical literals remain permitted where they identify historical records.
    offenders = []
    current_literal = re.compile(r"(?<![0-9])(?:v\.?)?" + re.escape(SEMVER) + r"(?![0-9])")
    for path in sorted(ROOT.rglob("*.py")):
        if ".git" in path.parts or path.name == "release_version.py":
            continue
        if current_literal.search(path.read_text(encoding="utf-8")):
            offenders.append(path.relative_to(ROOT).as_posix())
    ck("executable_python_does_not_hardcode_current_release", not offenders, offenders)

    harness_spec = yaml.safe_load((ROOT / "conformance/AGCP-Conformance-Harness-Spec.yml").read_text(encoding="utf-8")) or {}
    harness_checks = json.loads((ROOT / "conformance/harness-checks.json").read_text(encoding="utf-8"))
    ck(
        "conformance_harness_release_metadata",
        harness_spec.get("meta", {}).get("agcp_release") == RELEASE_TAG and harness_checks.get("agcp_release") == RELEASE_TAG,
        {"harness_spec": harness_spec.get("meta", {}).get("agcp_release"), "harness_checks": harness_checks.get("agcp_release")},
    )

    disposition = json.loads((ROOT / "governance/normative-companion-reference-dispositions.json").read_text(encoding="utf-8"))
    disposition_md = (ROOT / "governance/AGCP-Normative-Companion-Reference-Dispositions.md").read_text(encoding="utf-8")
    ck(
        "normative_companion_disposition_release_metadata",
        disposition.get("release_target") == RELEASE_TAG and f"**Release target:** AGCP {RELEASE_TAG}" in disposition_md,
        {"json_release_target": disposition.get("release_target")},
    )

    return checks, issues


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    changed, preissues = sync(args.write)
    checks, issues = validate()
    issues = preissues + issues
    report = {
        "release_context": release_context(),
        "validation_type": "AGCP_SINGLE_SOURCE_VERSION_VALIDATION",
        "status": "PASS" if not issues else "FAIL",
        "version_source": "VERSION",
        "semantic_version": SEMVER,
        "release_tag": RELEASE_TAG,
        "rtm_specification_version": RTM_SPEC_VERSION,
        "synchronization": {"files_requiring_update": changed},
        "checks": checks,
        "issues": issues,
    }
    if args.write or args.check:
        REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    print(json.dumps({"status": report["status"], "changed": changed, "issues": issues}, indent=2, default=str))
    return 0 if not issues else 1


if __name__ == "__main__":
    raise SystemExit(main())
